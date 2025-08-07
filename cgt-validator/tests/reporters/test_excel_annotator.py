"""Tests for Excel error annotator."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.reporters.excel_annotator import ExcelAnnotator
from src.reporters.validation_results import ValidationResults


class TestExcelAnnotator:
    """Test Excel annotation functionality."""

    def test_annotate_file_creates_output(
        self, invalid_oregon_excel: Path, sample_validation_results: ValidationResults
    ):
        """Test that annotation creates an output file."""
        annotator = ExcelAnnotator()

        # Annotate file
        output_path = annotator.annotate_file(str(invalid_oregon_excel), sample_validation_results)

        assert output_path.exists()
        assert "_annotated" in output_path.name
        assert output_path.suffix == ".xlsx"

    def test_annotate_file_custom_output(
        self, invalid_oregon_excel: Path, temp_dir: Path, sample_validation_results: ValidationResults
    ):
        """Test annotation with custom output path."""
        annotator = ExcelAnnotator()
        custom_output = temp_dir / "custom_annotated.xlsx"

        output_path = annotator.annotate_file(str(invalid_oregon_excel), sample_validation_results, str(custom_output))

        assert output_path == custom_output
        assert output_path.exists()

    def test_summary_sheet_added(self, invalid_oregon_excel: Path, sample_validation_results: ValidationResults):
        """Test that summary sheet is added."""
        annotator = ExcelAnnotator()
        output_path = annotator.annotate_file(str(invalid_oregon_excel), sample_validation_results)

        # Load annotated workbook
        wb = load_workbook(output_path)

        assert "Validation Summary" in wb.sheetnames
        assert wb.sheetnames[0] == "Validation Summary"  # Should be first sheet

        # Check summary content
        summary_sheet = wb["Validation Summary"]
        assert "CGT Validation Summary" in str(summary_sheet["A1"].value)
        assert summary_sheet["A3"].value == "State:"
        assert summary_sheet["B3"].value == "Oregon"

    def test_annotations_count(self, temp_dir: Path):
        """Test that annotations are counted correctly."""
        # Create Excel file matching the validation results
        wb = Workbook()

        # Create sheets that match the validation results
        ws1 = wb.active
        ws1.title = "Provider Information"
        ws1["A1"] = "Provider ID"
        ws1["B1"] = "NPI"
        ws1["A2"] = "PRV001"
        ws1["B2"] = "123"  # Invalid NPI

        ws2 = wb.create_sheet("Medical Claims")
        ws2["A1"] = "Claim ID"
        ws2["B1"] = "Paid Amount"
        ws2["A2"] = "CLM001"
        ws2["B2"] = -100  # Negative amount

        test_file = temp_dir / "test.xlsx"
        wb.save(test_file)

        # Create validation results matching the file
        results = ValidationResults("test", 2024)
        results.add_error("INVALID_NPI", "NPI must be 10 digits", "Provider Information.NPI")
        results.add_warning("NEGATIVE_AMOUNT", "Negative amounts found", "Medical Claims.Paid Amount")

        annotator = ExcelAnnotator()
        annotator.annotate_file(str(test_file), results)

        assert annotator.get_annotations_count() > 0

    def test_severity_styling(self, temp_dir: Path):
        """Test that different severities get different styling."""
        # Create a simple Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Column1"
        ws["A2"] = "Value1"

        test_file = temp_dir / "test.xlsx"
        wb.save(test_file)

        # Create validation results with different severities
        results = ValidationResults("test", 2024)
        results.add_error("ERROR_TEST", "Error in column", "TestSheet.Column1")
        results.add_warning("WARNING_TEST", "Warning in column", "TestSheet.Column1")

        # Annotate
        annotator = ExcelAnnotator()
        output_path = annotator.annotate_file(str(test_file), results)

        # Load and check styling
        wb_annotated = load_workbook(output_path)
        ws_annotated = wb_annotated["TestSheet"]

        # Header should have styling applied
        header_cell = ws_annotated["A1"]
        assert header_cell.fill.start_color.rgb is not None  # Has fill color
        assert header_cell.comment is not None  # Has comment

    def test_sheet_level_annotation(self, temp_dir: Path):
        """Test sheet-level annotations."""
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        test_file = temp_dir / "test.xlsx"
        wb.save(test_file)

        # Create validation result with sheet-level error
        results = ValidationResults("test", 2024)
        results.add_error("SHEET_ERROR", "Sheet-level error", "TestSheet")

        # Annotate
        annotator = ExcelAnnotator()
        output_path = annotator.annotate_file(str(test_file), results)

        # Check annotation
        wb_annotated = load_workbook(output_path)
        ws_annotated = wb_annotated["TestSheet"]

        # A1 should have annotation for sheet-level issues
        assert ws_annotated["A1"].comment is not None
        assert "SHEET_ERROR" in ws_annotated["A1"].comment.text

    def test_column_annotation_with_rows(self, temp_dir: Path):
        """Test column annotation with specific row references."""
        # Create Excel file with data
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        ws["A1"] = "ID"
        ws["B1"] = "Value"

        for i in range(2, 10):
            ws[f"A{i}"] = f"ID{i}"
            ws[f"B{i}"] = i * 10

        test_file = temp_dir / "test.xlsx"
        wb.save(test_file)

        # Create validation result mentioning specific rows
        results = ValidationResults("test", 2024)
        results.add_error("INVALID_VALUES", "Invalid values in rows: [3, 5, 7]", "DataSheet.Value")

        # Annotate
        annotator = ExcelAnnotator()
        output_path = annotator.annotate_file(str(test_file), results)

        # Check annotations
        wb_annotated = load_workbook(output_path)
        ws_annotated = wb_annotated["DataSheet"]

        # Column header should be annotated
        assert ws_annotated["B1"].comment is not None
        assert ws_annotated["B1"].fill.start_color.rgb is not None

        # Specific rows should be highlighted
        # Note: Row 3 in error message means row 4 in Excel (1-based + header)
        for excel_row in [4, 6, 8]:  # Rows 3, 5, 7 + header offset
            cell = ws_annotated[f"B{excel_row}"]
            assert cell.fill.start_color.rgb is not None

    def test_multiple_issues_same_location(self, temp_dir: Path):
        """Test handling multiple issues at the same location."""
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "TestColumn"

        test_file = temp_dir / "test.xlsx"
        wb.save(test_file)

        # Create multiple issues for same location
        results = ValidationResults("test", 2024)
        results.add_error("ERROR1", "First error", "TestSheet.TestColumn")
        results.add_error("ERROR2", "Second error", "TestSheet.TestColumn")
        results.add_warning("WARNING1", "First warning", "TestSheet.TestColumn")

        # Annotate
        annotator = ExcelAnnotator()
        output_path = annotator.annotate_file(str(test_file), results)

        # Check annotations
        wb_annotated = load_workbook(output_path)
        ws_annotated = wb_annotated["TestSheet"]

        # Should have comment with all issues
        comment = ws_annotated["A1"].comment
        assert comment is not None
        assert "ERROR1" in comment.text
        assert "ERROR2" in comment.text
        assert "WARNING1" in comment.text

        # Should use error styling (highest severity)
        # Check color - openpyxl may include alpha channel
        color = ws_annotated["A1"].fill.start_color.rgb
        assert color in ["FFFFE6E6", "00FFE6E6"]  # Error color (with or without alpha)

    def test_file_level_issues(self, temp_dir: Path):
        """Test handling of file-level issues."""
        # Create Excel file
        wb = Workbook()
        test_file = temp_dir / "test.xlsx"
        wb.save(test_file)

        # Create file-level issue
        results = ValidationResults("test", 2024)
        results.add_error("FILE_ERROR", "File-level error", "file")

        # Annotate
        annotator = ExcelAnnotator()
        output_path = annotator.annotate_file(str(test_file), results)

        # Should still create output with summary
        assert output_path.exists()

        wb_annotated = load_workbook(output_path)
        assert "Validation Summary" in wb_annotated.sheetnames
