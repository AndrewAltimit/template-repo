"""Shared pytest fixtures for CGT validator tests."""

import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, Generator

import pandas as pd
import pytest
from openpyxl import Workbook

from src.mock_data.oregon_generator import OregonMockDataGenerator
from src.reporters.validation_results import ValidationResults


@pytest.fixture
def temp_dir() -> Generator[Path, None, None]:
    """Create a temporary directory for test files."""
    temp_path = Path(tempfile.mkdtemp())
    yield temp_path
    shutil.rmtree(temp_path)


@pytest.fixture
def mock_oregon_data() -> OregonMockDataGenerator:
    """Create an Oregon mock data generator."""
    return OregonMockDataGenerator(seed=12345)


@pytest.fixture
def valid_oregon_excel(temp_dir: Path, mock_oregon_data: OregonMockDataGenerator) -> Path:
    """Create a valid Oregon submission Excel file."""
    output_path = temp_dir / "valid_oregon_submission.xlsx"
    mock_oregon_data.save_to_excel(str(output_path), include_validations=True)
    return output_path


@pytest.fixture
def invalid_oregon_excel(temp_dir: Path) -> Path:
    """Create an invalid Oregon submission Excel file with known issues."""
    output_path = temp_dir / "invalid_oregon_submission.xlsx"

    # Create workbook with issues
    wb = Workbook()

    # Provider Information - missing required fields
    ws_provider = wb.active
    ws_provider.title = "Provider Information"
    ws_provider.append(["Provider ID", "Provider Name", "Provider Type", "Tax ID"])  # Missing NPI, Address, etc.
    ws_provider.append(["PRV001", "Test Provider", "Hospital", "12-3456789"])
    ws_provider.append(["PRV002", "Test Provider 2", "Invalid Type", "invalid-tax"])  # Invalid type and tax ID

    # Member Months - wrong format
    ws_member = wb.create_sheet("Member Months")
    ws_member.append(["Provider ID", "Reporting Period", "Line of Business", "Member Months", "Unique Members"])
    ws_member.append(["PRV001", "2024/01", "Commercial", "abc", 100])  # Wrong date format, non-numeric member months
    ws_member.append(["PRV999", "2024-02", "Invalid LOB", 200, 250])  # Invalid provider ref, invalid LOB, unique > total

    # Medical Claims - missing sheet
    # (Intentionally not creating Medical Claims sheet)

    # Pharmacy Claims - data issues
    ws_pharmacy = wb.create_sheet("Pharmacy Claims")
    ws_pharmacy.append(
        [
            "Provider ID",
            "Member ID",
            "Claim ID",
            "Fill Date",
            "NDC",
            "Days Supply",
            "Quantity",
            "Allowed Amount",
            "Paid Amount",
            "Member Liability",
        ]
    )
    ws_pharmacy.append(
        ["PRV001", "MEM001", "RX001", "2024-01-15", "12345", -30, 100, 50.00, 60.00, -10.00]
    )  # Negative days, paid > allowed, negative liability

    # Save file
    wb.save(output_path)
    return output_path


@pytest.fixture
def missing_sheets_excel(temp_dir: Path) -> Path:
    """Create an Excel file missing required sheets."""
    output_path = temp_dir / "missing_sheets.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Provider Information"
    ws.append(["Provider ID", "Provider Name"])
    ws.append(["PRV001", "Test Provider"])

    # Missing Member Months, Medical Claims, Pharmacy Claims, Reconciliation
    wb.save(output_path)
    return output_path


@pytest.fixture
def empty_excel(temp_dir: Path) -> Path:
    """Create an empty Excel file."""
    output_path = temp_dir / "empty.xlsx"
    wb = Workbook()
    wb.save(output_path)
    return output_path


@pytest.fixture
def sample_validation_results() -> ValidationResults:
    """Create sample validation results for testing reporters."""
    results = ValidationResults(state="oregon", year=2024)

    # Add various issues
    results.add_error(code="MISSING_SHEET", message="Required sheet 'Medical Claims' not found", location="sheets")

    results.add_error(
        code="INVALID_NPI", message="NPI must be 10 digits in rows: [2, 5, 8]", location="Provider Information.NPI"
    )

    results.add_warning(
        code="NEGATIVE_AMOUNT",
        message="Negative amounts found in 'Paid Amount' rows: [15, 23]",
        location="Medical Claims.Paid Amount",
    )

    results.add_info(code="BEHAVIORAL_HEALTH_FOUND", message="Found 156 behavioral health claims", location="Medical Claims")

    return results


@pytest.fixture
def mock_state_config() -> Dict:
    """Create mock state configuration for testing."""
    return {
        "direct_urls": [
            {
                "url": "https://example.com/test.pdf",
                "type": "pdf",
                "description": "Test Document",
                "version": "1.0",
                "last_checked": datetime.now().strftime("%Y-%m-%d"),
            }
        ],
        "index_urls": [
            {
                "url": "https://example.com/index",
                "scan_pattern": r"\.pdf$",
                "keywords": ["template", "test"],
                "last_scraped": datetime.now().strftime("%Y-%m-%d"),
            }
        ],
    }


@pytest.fixture
def mock_web_response() -> str:
    """Create mock HTML response for web scraping tests."""
    return """
    <html>
    <body>
        <a href="/documents/template_v2.pdf">Template Version 2</a>
        <a href="/documents/manual_2024.xlsx">2024 Manual</a>
        <a href="/documents/old_doc.pdf">Old Document</a>
        <a href="/unrelated.html">Unrelated Page</a>
    </body>
    </html>
    """


@pytest.fixture
def csv_file(temp_dir: Path) -> Path:
    """Create a CSV file (invalid format for testing)."""
    output_path = temp_dir / "data.csv"
    df = pd.DataFrame({"Provider ID": ["PRV001", "PRV002"], "Provider Name": ["Test 1", "Test 2"]})
    df.to_csv(output_path, index=False)
    return output_path
