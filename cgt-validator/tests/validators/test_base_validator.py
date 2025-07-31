"""Tests for base validator functionality."""

from pathlib import Path
from typing import Any, Dict

import pandas as pd

from src.reporters.validation_results import ValidationResults
from src.validators.base_validator import ValidatorBase


class MockValidator(ValidatorBase):
    """Mock validator for testing base functionality."""

    def _load_requirements(self) -> Dict[str, Any]:
        return {
            "min_sheets": 2,
            "required_sheets": ["Sheet1", "Sheet2"],
            "all_sheets": ["Sheet1", "Sheet2", "Sheet3"],
            "mandatory_fields": {"Sheet1": ["ID", "Name", "Value"], "Sheet2": ["RefID", "Amount"]},
            "data_types": {"Sheet1": {"ID": "text", "Value": "numeric"}, "Sheet2": {"Amount": "numeric"}},
        }

    def _validate_business_rules(self, excel_data: pd.ExcelFile, results: ValidationResults):
        pass

    def _validate_cross_references(self, excel_data: pd.ExcelFile, results: ValidationResults):
        pass

    def _run_state_specific_validations(self, excel_data: pd.ExcelFile, results: ValidationResults):
        pass


class TestBaseValidator:
    """Test base validator functionality."""

    def test_check_data_type(self):
        """Test data type checking."""
        validator = MockValidator("test", 2024)

        # Numeric checks
        assert validator._check_data_type(123, "numeric")
        assert validator._check_data_type(123.45, "numeric")
        assert validator._check_data_type("123", "numeric")
        assert not validator._check_data_type("abc", "numeric")
        assert not validator._check_data_type("12.34.56", "numeric")

        # Text checks
        assert validator._check_data_type("hello", "text")
        assert not validator._check_data_type(123, "text")

        # Integer checks
        assert validator._check_data_type(123, "integer")
        assert validator._check_data_type(123.0, "integer")
        assert not validator._check_data_type(123.45, "integer")
        assert not validator._check_data_type("abc", "integer")

        # Date checks (requires datetime objects)
        from datetime import datetime

        assert validator._check_data_type(datetime.now(), "date")
        assert validator._check_data_type(pd.Timestamp.now(), "date")
        assert not validator._check_data_type("2024-01-01", "date")

    def test_validate_numeric_range(self, temp_dir: Path):
        """Test numeric range validation."""
        validator = MockValidator("test", 2024)
        results = ValidationResults("test", 2024)

        # Create test dataframe
        df = pd.DataFrame({"Value": [10, 20, 30, -5, 150]})

        # Test min validation
        validator.validate_numeric_range(df, "Value", 0, None, "TestSheet", results)
        assert any(e.code == "VALUE_TOO_LOW" for e in results.errors)

        # Test max validation
        results = ValidationResults("test", 2024)
        validator.validate_numeric_range(df, "Value", None, 100, "TestSheet", results)
        assert any(e.code == "VALUE_TOO_HIGH" for e in results.errors)

        # Test both min and max
        results = ValidationResults("test", 2024)
        validator.validate_numeric_range(df, "Value", 0, 100, "TestSheet", results)
        assert any(e.code == "VALUE_TOO_LOW" for e in results.errors)
        assert any(e.code == "VALUE_TOO_HIGH" for e in results.errors)

    def test_validate_allowed_values(self):
        """Test allowed values validation."""
        validator = MockValidator("test", 2024)
        results = ValidationResults("test", 2024)

        df = pd.DataFrame({"Status": ["Active", "Inactive", "Pending", "Invalid", "Active"]})

        allowed = ["Active", "Inactive", "Pending"]
        validator.validate_allowed_values(df, "Status", allowed, "TestSheet", results)

        assert any(e.code == "INVALID_VALUE" for e in results.errors)
        assert "Invalid" in results.errors[0].message

    def test_validate_unique_values(self):
        """Test unique values validation."""
        validator = MockValidator("test", 2024)
        results = ValidationResults("test", 2024)

        df = pd.DataFrame({"ID": ["001", "002", "003", "002", "004", "003"]})  # Duplicates: 002, 003

        validator.validate_unique_values(df, "ID", "TestSheet", results)

        assert any(e.code == "DUPLICATE_VALUES" for e in results.errors)
        dup_error = next(e for e in results.errors if e.code == "DUPLICATE_VALUES")
        assert "duplicate values" in dup_error.message

    def test_file_access_validation(self, temp_dir: Path):
        """Test file access validation."""
        from openpyxl import Workbook

        validator = MockValidator("test", 2024)

        # Test non-existent file
        results = ValidationResults("test", 2024)
        assert not validator._check_file_access(Path("/non/existent.xlsx"), results)
        assert any(e.code == "FILE_NOT_FOUND" for e in results.errors)

        # Test wrong file type
        txt_file = temp_dir / "test.txt"
        txt_file.write_text("test")
        results = ValidationResults("test", 2024)
        assert not validator._check_file_access(txt_file, results)
        assert any(e.code == "INVALID_FILE_TYPE" for e in results.errors)

        # Test empty file
        empty_file = temp_dir / "empty.xlsx"
        empty_file.write_bytes(b"")
        results = ValidationResults("test", 2024)
        assert not validator._check_file_access(empty_file, results)
        assert any(e.code == "EMPTY_FILE" for e in results.errors)

        # Test valid file
        valid_file = temp_dir / "valid.xlsx"
        wb = Workbook()
        wb.save(valid_file)
        results = ValidationResults("test", 2024)
        assert validator._check_file_access(valid_file, results)
        assert len(results.errors) == 0

    def test_sheet_validation(self, temp_dir: Path):
        """Test sheet structure validation."""
        from openpyxl import Workbook

        validator = MockValidator("test", 2024)

        # Create file with wrong sheets
        output_path = temp_dir / "sheet_test.xlsx"
        wb = Workbook()

        # Only one sheet (requires 2)
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["ID", "Name", "Value"])
        ws.append(["001", "Test", 100])

        # Add unexpected sheet
        ws = wb.create_sheet("UnexpectedSheet")
        ws.append(["Data"])

        wb.save(output_path)

        results = validator.validate_file(str(output_path))

        # Should have insufficient sheets error
        assert any(e.code == "INSUFFICIENT_SHEETS" for e in results.errors)

        # Should have missing sheet error for Sheet2
        assert any(e.code == "MISSING_SHEET" and "Sheet2" in e.message for e in results.errors)

        # Should have warning about unexpected sheet
        assert any(w.code == "UNEXPECTED_SHEET" and "UnexpectedSheet" in w.message for w in results.warnings)

    def test_mandatory_fields_validation(self, temp_dir: Path):
        """Test mandatory fields validation."""
        from openpyxl import Workbook

        validator = MockValidator("test", 2024)

        output_path = temp_dir / "mandatory_test.xlsx"
        wb = Workbook()

        # Sheet1 with missing mandatory column
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["ID", "Name"])  # Missing "Value"
        ws.append(["001", "Test"])

        # Sheet2 with empty mandatory values
        ws = wb.create_sheet("Sheet2")
        ws.append(["RefID", "Amount"])
        ws.append(["", 100])  # Empty RefID
        ws.append(["002", ""])  # Empty Amount

        wb.save(output_path)

        results = validator.validate_file(str(output_path))

        # Should have missing column error
        assert any(e.code == "MISSING_COLUMN" and "Value" in e.message for e in results.errors)

        # Should have empty mandatory field errors
        assert any(e.code == "EMPTY_MANDATORY_FIELD" and "RefID" in e.message for e in results.errors)
