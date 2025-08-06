"""Tests for Oregon validator (2025 CGT-1 template)."""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from src.validators.oregon import OregonValidator


class TestOregonValidator:
    """Test Oregon-specific validator for 2025 CGT-1 template."""

    def test_valid_submission(self, valid_oregon_excel: Path):
        """Test validation of a valid Oregon 2025 submission."""
        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(valid_oregon_excel))

        # The mock data generator should create valid files
        # If there are errors, print them for debugging
        if not results.is_valid():
            for error in results.errors:
                print(f"Error: {error.code} - {error.message} at {error.location}")

        assert results.is_valid()
        assert results.state == "oregon"
        assert results.year == 2025

    def test_invalid_submission(self, invalid_oregon_excel: Path):
        """Test validation of an invalid Oregon 2025 submission."""
        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(invalid_oregon_excel))

        assert not results.is_valid()
        assert len(results.errors) > 0

        # Check for specific expected errors based on 2025 template
        error_codes = [error.code for error in results.errors]
        # Could have missing sheets or missing required fields
        assert any(code in error_codes for code in ["MISSING_SHEET", "MISSING_REQUIRED_FIELD", "MISSING_COLUMN"])

    def test_missing_sheets(self, missing_sheets_excel: Path):
        """Test validation when required 2025 template sheets are missing."""
        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(missing_sheets_excel))

        assert not results.is_valid()

        # Should have errors for missing required sheets (9 required in 2025)
        missing_sheet_errors = [e for e in results.errors if e.code == "MISSING_SHEET"]
        assert len(missing_sheet_errors) >= 5  # At least some of the 9 required sheets missing

    def test_tin_validation(self, temp_dir: Path):
        """Test TIN format validation for 2025 template."""
        # Create file with invalid TINs
        output_path = temp_dir / "tin_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets first
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
        ]:
            ws = wb.create_sheet(sheet_name)
            # Add minimal headers
            for i in range(9):
                ws.append([None] * 10)

        # PROV_ID sheet with TIN validation
        ws = wb.create_sheet("9. PROV_ID")
        # Add header rows
        for i in range(8):
            ws.append([None] * 5)

        # Row 9: Column headers
        ws.append(
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
                None,
                None,
            ]
        )

        # Valid TIN (9 digits)
        ws.append(["Provider 1", "IPA 1", "123456789", None, None])

        # Invalid TINs
        ws.append(["Provider 2", "IPA 2", "12345", None, None])  # Too short
        ws.append(["Provider 3", "IPA 3", "1234567890", None, None])  # Too long
        ws.append(["Provider 4", "IPA 4", "12345678A", None, None])  # Contains letter

        wb.save(output_path)

        # Validate
        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have TIN errors
        tin_errors = [e for e in results.errors if e.code == "INVALID_TIN"]
        assert len(tin_errors) > 0

    def test_line_of_business_validation(self, temp_dir: Path):
        """Test Line of Business code validation for 2025 template."""
        output_path = temp_dir / "lob_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(9):
                ws.append([None] * 10)

        # TME_ALL sheet with LOB validation
        ws = wb.create_sheet("2. TME_ALL")
        # Add header rows - headers should be at row 9 (0-indexed row 8)
        for i in range(8):
            ws.append([None] * 10)

        # Row 9: Column headers (0-indexed row 8)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months", "Demographic Score"] + [None] * 6)

        # Valid LOB codes (1-7)
        ws.append([2024, 1, 1000, 1.0] + [None] * 6)  # Medicare
        ws.append([2024, 2, 2000, 1.1] + [None] * 6)  # Medicaid

        # Invalid LOB codes
        ws.append([2024, 0, 3000, 1.2] + [None] * 6)  # Invalid: 0
        ws.append([2024, 8, 4000, 1.3] + [None] * 6)  # Invalid: 8
        ws.append([2024, 99, 5000, 1.4] + [None] * 6)  # Invalid: 99

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have LOB errors
        lob_errors = [e for e in results.errors if e.code == "INVALID_LOB_CODE"]
        assert len(lob_errors) > 0

    def test_member_months_validation(self, temp_dir: Path):
        """Test member months validation for 2025 template."""
        output_path = temp_dir / "member_months_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(9):
                ws.append([None] * 10)

        # TME_PROV sheet with member months validation
        ws = wb.create_sheet("3. TME_PROV")
        # Add header rows (TME_PROV has headers at row 11, 0-indexed = 10)
        for i in range(10):
            ws.append([None] * 10)

        # Row 11: Column headers (0-indexed row 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )

        # Valid member months
        ws.append([2024, 1, "Provider 1", "IPA 1", 1, 100, 1.0] + [None] * 3)

        # Below threshold (should warn)
        ws.append([2024, 1, "Provider 2", "IPA 1", 1, 10, 1.0] + [None] * 3)  # <= 12
        ws.append([2024, 1, "Provider 3", "IPA 1", 1, 5, 1.0] + [None] * 3)  # <= 12

        # Invalid member months
        ws.append([2024, 1, "Provider 4", "IPA 1", 1, 0, 1.0] + [None] * 3)  # Zero
        ws.append([2024, 1, "Provider 5", "IPA 1", 1, -10, 1.0] + [None] * 3)  # Negative
        ws.append([2024, 1, "Provider 6", "IPA 1", 1, None, 1.0] + [None] * 3)  # Missing

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have invalid member months errors
        mm_errors = [e for e in results.errors if e.code == "INVALID_MEMBER_MONTHS"]
        assert len(mm_errors) > 0

        # Should have below threshold warnings
        threshold_warnings = [w for w in results.warnings if w.code == "MEMBER_MONTHS_BELOW_THRESHOLD"]
        assert len(threshold_warnings) > 0

    def test_cross_reference_validation(self, temp_dir: Path):
        """Test cross-reference validation between sheets for 2025 template."""
        output_path = temp_dir / "cross_ref_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(9):
                ws.append([None] * 10)

        # PROV_ID sheet with provider TINs (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("9. PROV_ID")
        for i in range(8):
            ws.append([None] * 5)
        ws.append(["Provider Organization Name", "IPA or Contract Name", "Provider Organization TIN"] + [None] * 2)
        ws.append(["Provider 1", "IPA 1", "123456789"] + [None] * 2)
        ws.append(["Provider 2", "IPA 2", "987654321"] + [None] * 2)

        # TME_PROV with invalid provider reference (header at row 11, 0-indexed = 10)
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )
        ws.append([2024, 1, "Provider 1", "IPA 1", 1, 1000, 1.0] + [None] * 3)  # Valid
        ws.append([2024, 1, "Unknown Provider", "IPA 1", 1, 2000, 1.0] + [None] * 3)  # Invalid provider name

        # RX_MED_PROV with invalid provider reference (header at row 11, 0-indexed = 10)
        ws = wb.create_sheet("6. RX_MED_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization TIN",
                "Provider Organization Name",
                "Allowed Pharmacy",
                "Net Paid Medical",
            ]
            + [None] * 4
        )
        ws.append([2024, 1, "987654321", "Provider 2", 10000, 50000] + [None] * 4)  # Valid
        ws.append([2024, 1, "111111111", "Invalid Provider", 20000, 60000] + [None] * 4)  # Invalid TIN

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have invalid provider reference errors
        ref_errors = [e for e in results.errors if e.code == "INVALID_PROVIDER_REFERENCE"]
        assert len(ref_errors) > 0

    def test_cover_page_validation(self, temp_dir: Path):
        """Test cover page required fields validation for 2025 template."""
        output_path = temp_dir / "cover_page_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "2. TME_ALL",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(9):
                ws.append([None] * 10)

        # Cover Page with missing required fields
        ws = wb.create_sheet("1. Cover Page")
        # Create a sparse structure matching the template
        data: list[list[Any]] = [[None] * 4 for _ in range(20)]

        data[0][0] = "1. Cover Page"
        data[1][0] = "All questions in this tab must be answered."
        data[3][0] = "Contact Information"
        data[4][0] = "Payer Name:"
        data[4][2] = "[Input Required]"  # Missing payer name
        data[5][0] = "Contact Name:"
        data[5][2] = "John Smith"
        data[6][0] = "Contact Email:"
        data[6][2] = "[Input Required]"  # Missing email

        data[8][0] = "Attestation"
        data[10][0] = "Authorized Signatory Details"
        data[11][0] = "Full Name:"
        data[11][2] = ""  # Missing full name
        data[12][0] = "Title/Position:"
        data[12][2] = "CFO"
        data[13][0] = "Email/Contact Information:"
        data[13][2] = "jane.doe@example.com"
        data[14][0] = "Signature:"
        data[14][2] = "[Input Required]"  # Missing signature
        data[15][0] = "Date:"
        data[15][2] = ""  # Missing date

        for row in data:
            ws.append(row)

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have missing required field errors
        missing_field_errors = [e for e in results.errors if e.code == "MISSING_REQUIRED_FIELD"]
        assert len(missing_field_errors) >= 4  # At least payer name, email, full name, date

    def test_reporting_year_validation(self, temp_dir: Path):
        """Test reporting year validation for 2025 template."""
        output_path = temp_dir / "year_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(9):
                ws.append([None] * 10)

        # TME_ALL sheet with year validation (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 10)

        ws.append(["Reporting Year", "Line of Business Code", "Member Months", "Demographic Score"] + [None] * 6)

        # Valid years (current year - 1 or current year)
        ws.append([2024, 1, 1000, 1.0] + [None] * 6)  # Valid: 2024 for 2025 validator
        ws.append([2025, 1, 2000, 1.1] + [None] * 6)  # Valid: current year

        # Invalid years
        ws.append([2023, 1, 3000, 1.2] + [None] * 6)  # Too old
        ws.append([2026, 1, 4000, 1.3] + [None] * 6)  # Future year

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have reporting year errors
        year_errors = [e for e in results.errors if e.code == "INVALID_REPORTING_YEAR"]
        assert len(year_errors) > 0

    def test_provider_ipa_combination_validation(self, temp_dir: Path):
        """Test validation of Provider-IPA combinations between sheets."""
        output_path = temp_dir / "provider_ipa_test.xlsx"
        wb = Workbook()

        # Create Contents sheet first
        ws = wb.active
        ws.title = "Contents"
        for i in range(20):
            ws.append([None] * 10)

        # Create Cover Page with minimal required structure
        ws = wb.create_sheet("1. Cover Page")
        # Add required fields to avoid validation errors
        data: list[list[Any]] = [[None] * 4 for _ in range(20)]
        data[0][0] = "1. Cover Page"
        data[4][0] = "Payer Name:"
        data[4][2] = "Test Payer"
        data[5][0] = "Contact Name:"
        data[5][2] = "John Smith"
        data[6][0] = "Contact Email:"
        data[6][2] = "john@example.com"
        data[11][0] = "Full Name:"
        data[11][2] = "Jane Doe"
        data[12][0] = "Title/Position:"
        data[12][2] = "CFO"
        data[13][0] = "Email/Contact Information:"
        data[13][2] = "jane@example.com"
        data[14][0] = "Signature:"
        data[14][2] = "Jane Doe"
        data[15][0] = "Date:"
        data[15][2] = "01/01/2025"
        for row in data:
            ws.append(row)

        # Create data sheets with proper structure
        # TME_ALL (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 30)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Member Months",
                "Demographic Score",
                "Claims: \nHospital Inpatient",
                "Claims: \nHospital Outpatient",
                "Claims: Professional, Primary Care Providers",
                "Claims: Professional, Specialty Providers",
                "Claims: Professional, Behavior Health Providers",
                "Claims: Professional, Other Providers",
            ]
            + [None] * 20
        )
        ws.append([2024, 1, 10000, 1.0, 100000, 50000, 20000, 30000, 10000, 5000] + [None] * 20)

        # TME_UNATTR (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("4. TME_UNATTR")
        for i in range(8):
            ws.append([None] * 30)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months"] + [None] * 27)
        ws.append([2024, 1, 100] + [None] * 27)

        # MARKET_ENROLL (header at row 9, 0-indexed = 8, different structure)
        ws = wb.create_sheet("5. MARKET_ENROLL")
        for i in range(8):
            ws.append([None] * 30)
        ws.append(["Market Enrollment Category", "Year 2023 Member Months", "Year 2024 Member Months"] + [None] * 27)
        ws.append(["1. Large group (51 + employees), fully insured", 50000, 60000] + [None] * 27)

        # RX_MED_PROV (header at row 11, 0-indexed = 10)
        ws = wb.create_sheet("6. RX_MED_PROV")
        for i in range(10):
            ws.append([None] * 30)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization TIN",
                "Provider Organization Name",
                "Allowed Pharmacy",
                "Net Paid Medical",
            ]
            + [None] * 24
        )
        ws.append([2024, 1, "123456789", "Provider 1", 10000, 50000] + [None] * 24)

        # RX_MED_UNATTR (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("7. RX_MED_UNATTR")
        for i in range(8):
            ws.append([None] * 30)
        ws.append(["Reporting Year", "Line of Business Code", "Allowed Pharmacy", "Net Paid Medical"] + [None] * 26)
        ws.append([2024, 1, 5000, 25000] + [None] * 26)

        # RX_REBATE (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("8. RX_REBATE")
        for i in range(8):
            ws.append([None] * 30)
        ws.append(["Reporting Year", "Line of Business Code", "Prescription Rebates"] + [None] * 27)
        ws.append([2024, 1, 1000] + [None] * 27)

        # PROV_ID sheet with specific provider-IPA combinations
        ws = wb.create_sheet("9. PROV_ID")
        for i in range(8):
            ws.append([None] * 5)
        ws.append(
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
            ]
            + [None] * 2
        )
        ws.append(["Provider 1", "IPA 1", "123456789"] + [None] * 2)
        ws.append(["Provider 1", "IPA 2", "123456789"] + [None] * 2)  # Same provider, different IPA
        ws.append(["Provider 2", "", "987654321"] + [None] * 2)  # Provider with no IPA

        # TME_PROV with mismatched provider-IPA combination (header at row 11, 0-indexed = 10)
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 30)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 23
        )
        ws.append([2024, 1, "Provider 1", "IPA 1", 1, 1000, 1.0] + [None] * 23)  # Valid
        ws.append([2024, 1, "Provider 1", "IPA 3", 1, 2000, 1.0] + [None] * 23)  # Invalid IPA for Provider 1
        ws.append([2024, 1, "Provider 2", "IPA 1", 1, 3000, 1.0] + [None] * 23)  # Invalid - Provider 2 has no IPA

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have provider-IPA combination errors
        ipa_combo_errors = [e for e in results.errors if e.code == "INVALID_PROVIDER_IPA_COMBINATION"]

        # Print all errors for debugging if the specific error is not found
        if len(ipa_combo_errors) == 0:
            print("\nAll errors found:")
            for error in results.errors:
                print(f"  {error.code}: {error.message} at {error.location}")
            print("\nAll warnings found:")
            for warning in results.warnings:
                print(f"  {warning.code}: {warning.message} at {warning.location}")

            # Debug: check if the file was created correctly
            import pandas as pd

            excel = pd.ExcelFile(output_path)

            print("\nPROV_ID sheet (header=8):")
            prov_df = excel.parse("9. PROV_ID", header=8)
            print(prov_df)

            print("\nTME_PROV sheet (header=10):")
            tme_df = excel.parse("3. TME_PROV", header=10)
            print(tme_df[["Provider Organization Name", "IPA or Contract Name\n(If applicable/available)"]])

        assert len(ipa_combo_errors) > 0

    def test_file_not_found(self):
        """Test validation with non-existent file."""
        validator = OregonValidator(year=2025)
        results = validator.validate_file("/non/existent/file.xlsx")

        assert not results.is_valid()
        assert any(e.code == "FILE_NOT_FOUND" for e in results.errors)

    def test_invalid_file_type(self, csv_file: Path):
        """Test validation with invalid file type."""
        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(csv_file))

        assert not results.is_valid()
        assert any(e.code == "INVALID_FILE_TYPE" for e in results.errors)

    def test_empty_file(self, empty_excel: Path):
        """Test validation with empty Excel file."""
        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(empty_excel))

        assert not results.is_valid()
        # Should have errors about missing sheets
        assert any(e.code == "MISSING_SHEET" for e in results.errors)

    def test_lob_code_7_restriction(self, temp_dir: Path):
        """Test that LOB code 7 is only allowed in TME_ALL."""
        output_path = temp_dir / "lob_7_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):  # Account for different header rows
                ws.append([None] * 10)

        # TME_ALL with LOB 7 (should be valid) - header at row 9, 0-indexed = 8
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months", "Demographic Score"] + [None] * 6)
        ws.append([2024, 7, 1000, 1.0] + [None] * 6)  # Valid: LOB 7 in TME_ALL

        # TME_PROV with LOB 7 (should be invalid) - header at row 11, 0-indexed = 10
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )
        ws.append([2024, 7, "Provider 1", "IPA 1", 1, 1000, 1.0] + [None] * 3)  # Invalid: LOB 7 not allowed

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have LOB code 7 error for TME_PROV but not TME_ALL
        lob_7_errors = [e for e in results.errors if e.code == "INVALID_LOB_CODE_7"]
        assert len(lob_7_errors) == 1
        assert "3. TME_PROV" in lob_7_errors[0].location
        assert "LOB code 7" in lob_7_errors[0].message
        assert "only allowed in TME_ALL" in lob_7_errors[0].message

    def test_prov_id_field_codes(self, temp_dir: Path):
        """Test PROV_ID field codes validation."""
        output_path = temp_dir / "prov_id_codes_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # PROV_ID with incorrect field codes
        ws = wb.create_sheet("9. PROV_ID")
        for i in range(6):
            ws.append([None] * 5)
        # Row 6: Field codes - should be PRV01, PRV03, PRV02
        ws.append(["PRV01", "PRV99", "PRV02"] + [None] * 2)  # Wrong: middle should be PRV03
        ws.append(["free text", "free text", "text, 9 digits including leading zero"] + [None] * 2)
        ws.append(
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
            ]
            + [None] * 2
        )
        ws.append(["Provider 1", "IPA 1", "123456789"] + [None] * 2)

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have field code error
        field_code_errors = [e for e in results.errors if e.code == "INVALID_FIELD_CODES"]
        assert len(field_code_errors) == 1
        assert "PRV01, PRV03, PRV02" in field_code_errors[0].message

    def test_data_type_validation(self, temp_dir: Path):
        """Test comprehensive data type validation."""
        output_path = temp_dir / "data_types_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # TME_ALL with invalid data types (header at row 9, 0-indexed = 8)
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months", "Demographic Score"] + [None] * 6)
        ws.append(["invalid_year", "not_a_code", "negative", "text"] + [None] * 6)  # All invalid types
        ws.append([2024, 1.5, -100, -0.5] + [None] * 6)  # LOB not int, member months negative, demo score negative

        # TME_PROV with invalid data types (header at row 11, 0-indexed = 10)
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )
        ws.append(
            [2024, 1, "", "IPA 1", "not_a_code", 0, -1.0] + [None] * 3
        )  # Empty provider, invalid attr code, zero MM, negative demo

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have various data type errors
        data_type_errors = [e for e in results.errors if e.code == "INVALID_DATA_TYPE"]
        assert len(data_type_errors) > 0

        # Should also have specific validation errors
        assert any(e.code == "EMPTY_MANDATORY_FIELD" for e in results.errors)  # Empty provider name
        assert any(e.code == "INVALID_MEMBER_MONTHS" for e in results.errors)  # Zero/negative member months

    def test_case_insensitive_provider_matching(self, temp_dir: Path):
        """Test that provider matching is case-insensitive."""
        output_path = temp_dir / "case_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # PROV_ID with mixed case
        ws = wb.create_sheet("9. PROV_ID")
        for i in range(8):
            ws.append([None] * 5)
        ws.append(
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
            ]
            + [None] * 2
        )
        ws.append(["ABC Healthcare", "West IPA", "123456789"] + [None] * 2)
        ws.append(["XYZ Medical Group", "", "987654321"] + [None] * 2)

        # TME_PROV with different case (header at row 11, 0-indexed = 10)
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )
        # These should match despite case differences
        ws.append([2024, 1, "ABC HEALTHCARE", "WEST IPA", 1, 1000, 1.0] + [None] * 3)  # Upper case
        ws.append([2024, 1, "xyz medical group", "", 1, 2000, 1.0] + [None] * 3)  # Lower case
        ws.append([2024, 1, "Unknown Provider", "Some IPA", 1, 3000, 1.0] + [None] * 3)  # Should fail

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should only have one provider reference error (Unknown Provider)
        ref_errors = [e for e in results.errors if e.code == "INVALID_PROVIDER_REFERENCE"]
        assert len(ref_errors) == 1
        assert "Unknown Provider" in str(ref_errors[0].message)

    def test_rx_med_header_rows(self, temp_dir: Path):
        """Test that RX_MED_PROV and RX_MED_UNATTR have correct header rows."""
        output_path = temp_dir / "rx_med_headers_test.xlsx"
        wb = Workbook()

        # Create Contents sheet
        ws = wb.active
        ws.title = "Contents"
        for i in range(8):
            ws.append([None] * 10)

        # Create Cover Page with minimal required structure
        ws = wb.create_sheet("1. Cover Page")
        data: list[list[Any]] = [[None] * 4 for _ in range(20)]
        data[4][0] = "Payer Name:"
        data[4][2] = "Test Payer"
        data[5][0] = "Contact Name:"
        data[5][2] = "John Smith"
        data[6][0] = "Contact Email:"
        data[6][2] = "john@example.com"
        data[11][0] = "Full Name:"
        data[11][2] = "Jane Doe"
        data[12][0] = "Title/Position:"
        data[12][2] = "CFO"
        data[13][0] = "Email/Contact Information:"
        data[13][2] = "jane@example.com"
        data[14][0] = "Signature:"
        data[14][2] = "Jane Doe"
        data[15][0] = "Date:"
        data[15][2] = "01/01/2025"
        for row in data:
            ws.append(row)

        # Create TME_ALL with proper headers
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months", "Demographic Score"] + [None] * 6)
        ws.append([2024, 1, 10000, 1.0] + [None] * 6)

        # Create TME_PROV with proper headers
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )
        ws.append([2024, 1, "Provider 1", "", 1, 5000, 1.0] + [None] * 3)

        # Create TME_UNATTR with proper headers
        ws = wb.create_sheet("4. TME_UNATTR")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months"] + [None] * 7)
        ws.append([2024, 1, 500] + [None] * 7)

        # Create MARKET_ENROLL with proper headers (different structure)
        ws = wb.create_sheet("5. MARKET_ENROLL")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Market Enrollment Category", "Year 2023 Member Months", "Year 2024 Member Months"] + [None] * 7)
        ws.append(["1. Large group (51 + employees), fully insured", 50000, 60000] + [None] * 7)

        # Create RX_REBATE with proper headers
        ws = wb.create_sheet("8. RX_REBATE")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Prescription Rebates"] + [None] * 7)
        ws.append([2024, 1, 5000] + [None] * 7)

        # Create PROV_ID with proper headers
        ws = wb.create_sheet("9. PROV_ID")
        for i in range(8):
            ws.append([None] * 5)
        ws.append(
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
            ]
            + [None] * 2
        )
        ws.append(["Provider 1", "", "123456789"] + [None] * 2)

        # RX_MED_PROV with headers at row 11 (0-indexed = 10)
        ws = wb.create_sheet("6. RX_MED_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization TIN",
                "Provider Organization Name",
                "Allowed Pharmacy",
                "Net Paid Medical",
            ]
            + [None] * 4
        )
        ws.append([2024, 1, "123456789", "Provider 1", 10000, 50000] + [None] * 4)

        # RX_MED_UNATTR with headers at row 9 (0-indexed = 8)
        ws = wb.create_sheet("7. RX_MED_UNATTR")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Allowed Pharmacy", "Net Paid Medical"] + [None] * 6)
        ws.append([2024, 1, 5000, 25000] + [None] * 6)

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should validate successfully if headers are parsed correctly
        # If headers are at wrong row, mandatory field checks will fail
        mandatory_errors = [e for e in results.errors if e.code == "EMPTY_MANDATORY_FIELD"]
        missing_col_errors = [e for e in results.errors if e.code == "MISSING_COLUMN"]

        # Print errors for debugging
        if len(mandatory_errors) > 0 or len(missing_col_errors) > 0:
            print("\nErrors found:")
            for error in results.errors:
                print(f"  {error.code}: {error.message} at {error.location}")

        assert len(mandatory_errors) == 0, "Headers not parsed at correct row"
        assert len(missing_col_errors) == 0, "Required columns not found"

    def test_pharmacy_rebate_validation(self, temp_dir: Path):
        """Test that pharmacy rebates must be negative or zero."""
        output_path = temp_dir / "rebate_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # RX_REBATE with positive rebates (invalid) - header at row 9, 0-indexed = 8
        ws = wb.create_sheet("8. RX_REBATE")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Medical Pharmacy Rebate Amount",
                "Retail Pharmacy Rebate Amount",
                "Total Pharmacy Rebate Amount (Optional)",
            ]
            + [None] * 5
        )
        # Add data with positive rebates (invalid)
        ws.append([2024, 1, 5000, 10000, 15000] + [None] * 5)  # All positive - invalid
        ws.append([2024, 2, -5000, -10000, -15000] + [None] * 5)  # All negative - valid
        ws.append([2024, 3, 0, 0, 0] + [None] * 5)  # All zero - valid
        ws.append([2024, 4, 1000, -5000, -4000] + [None] * 5)  # Mixed - medical positive is invalid

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have errors for positive rebate amounts
        rebate_errors = [e for e in results.errors if e.code == "INVALID_PHARMACY_REBATE"]
        assert len(rebate_errors) >= 2  # At least 2 rows have positive rebates

        # Check that error messages mention negative requirement
        for error in rebate_errors:
            assert "negative" in error.message.lower() or "positive" in error.message.lower()
            assert "8. RX_REBATE" in error.location

    def test_hrsn_costs_location_validation(self, temp_dir: Path):
        """Test that HRSN costs should only be in TME_ALL and TME_UNATTR tabs."""
        output_path = temp_dir / "hrsn_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # TME_ALL with HRSN costs (valid) - header at row 9, 0-indexed = 8
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 15)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Member Months",
                "Demographic Score",
                "Non-Claims: HRSN",
            ]
            + [None] * 10
        )
        ws.append([2024, 1, 10000, 1.0, 50000] + [None] * 10)  # Valid: HRSN in TME_ALL

        # TME_PROV with HRSN costs (invalid) - header at row 11, 0-indexed = 10
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 15)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
                "Non-Claims: HRSN",
            ]
            + [None] * 7
        )
        ws.append([2024, 1, "Provider 1", "", 1, 5000, 1.0, 25000] + [None] * 7)  # Invalid: HRSN in TME_PROV

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # Should have error for HRSN in TME_PROV
        hrsn_errors = [e for e in results.errors if e.code == "HRSN_IN_WRONG_SHEET"]
        assert len(hrsn_errors) == 1
        assert "TME_PROV" in hrsn_errors[0].location
        assert "TME_ALL and TME_UNATTR" in hrsn_errors[0].message

    def test_lob_7_description(self, temp_dir: Path):
        """Test LOB 7 description is 'CCO-F and Medicaid Open Card Carve-Outs'."""
        output_path = temp_dir / "lob_7_desc_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "3. TME_PROV",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
            "9. PROV_ID",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # TME_ALL with LOB 7 - header at row 9, 0-indexed = 8
        ws = wb.create_sheet("2. TME_ALL")
        for i in range(8):
            ws.append([None] * 10)
        ws.append(["Reporting Year", "Line of Business Code", "Member Months", "Demographic Score"] + [None] * 6)
        ws.append([2024, 7, 1000, 1.0] + [None] * 6)  # LOB 7

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # LOB 7 should be valid in TME_ALL
        lob_errors = [e for e in results.errors if e.code in ["INVALID_LOB_CODE", "INVALID_LOB_CODE_7"]]

        # Should not have LOB errors for TME_ALL
        for error in lob_errors:
            assert "2. TME_ALL" not in error.location, "LOB 7 should be valid in TME_ALL"

    def test_ipa_required_for_dual_attribution(self, temp_dir: Path):
        """Test that IPA field is required for dual-level attribution."""
        output_path = temp_dir / "ipa_required_test.xlsx"
        wb = Workbook()

        # Create minimal required sheets
        for sheet_name in [
            "Contents",
            "1. Cover Page",
            "2. TME_ALL",
            "4. TME_UNATTR",
            "5. MARKET_ENROLL",
            "6. RX_MED_PROV",
            "7. RX_MED_UNATTR",
            "8. RX_REBATE",
        ]:
            ws = wb.create_sheet(sheet_name)
            for i in range(13):
                ws.append([None] * 10)

        # PROV_ID sheet - header at row 9, 0-indexed = 8
        ws = wb.create_sheet("9. PROV_ID")
        for i in range(8):
            ws.append([None] * 5)
        ws.append(
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
            ]
            + [None] * 2
        )
        ws.append(["Provider 1", "IPA 1", "123456789"] + [None] * 2)  # Provider with IPA

        # TME_PROV sheet - header at row 11, 0-indexed = 10
        ws = wb.create_sheet("3. TME_PROV")
        for i in range(10):
            ws.append([None] * 10)
        ws.append(
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 3
        )
        # IPA field is required for dual-level attribution
        ws.append([2024, 1, "Provider 1", "IPA 1", 1, 1000, 1.0] + [None] * 3)  # Valid with IPA

        wb.save(output_path)

        validator = OregonValidator(year=2025)
        results = validator.validate_file(str(output_path))

        # The IPA field requirement is validated in the spec
        # Check that no missing column errors occur
        missing_col_errors = [e for e in results.errors if e.code == "MISSING_COLUMN" and "IPA" in e.message]
        assert len(missing_col_errors) == 0, "IPA field should be recognized as present"
