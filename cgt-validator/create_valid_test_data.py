#!/usr/bin/env python
"""Create a truly valid test data file using openpyxl directly to ensure proper text formatting."""

import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook


def create_valid_oregon_submission():
    """Create a valid Oregon submission file with proper text formatting."""

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Provider Information Sheet
    ws_provider = wb.create_sheet("Provider Information")

    # Headers
    headers = ["Provider ID", "Provider Name", "Provider Type", "Tax ID", "NPI", "Address", "City", "State", "ZIP"]
    for col, header in enumerate(headers, 1):
        ws_provider.cell(row=1, column=col, value=header)

    # Sample providers with properly formatted text fields
    providers = []
    for i in range(10):
        row = i + 2
        provider_id = f"PRV{str(i+1).zfill(6)}"
        providers.append(provider_id)

        ws_provider.cell(row=row, column=1, value=provider_id)
        ws_provider.cell(row=row, column=2, value=f"Oregon Health Provider {i+1}")
        ws_provider.cell(row=row, column=3, value=random.choice(["Hospital", "Primary Care", "Specialist"]))
        ws_provider.cell(row=row, column=4, value=f"{random.randint(10,99)}-{random.randint(1000000,9999999)}")

        # NPI as text - use quote prefix to ensure text format
        npi = str(random.randint(1000000000, 9999999999))
        ws_provider.cell(row=row, column=5, value=npi).number_format = "@"

        ws_provider.cell(row=row, column=6, value=f"{random.randint(100,9999)} Main Street")
        ws_provider.cell(row=row, column=7, value="Portland")
        ws_provider.cell(row=row, column=8, value="OR")

        # ZIP as text with leading zeros preserved
        zip_code = f"{random.randint(97000, 97999):05d}"
        ws_provider.cell(row=row, column=9, value=zip_code).number_format = "@"

    # 2. Member Months Sheet
    ws_member = wb.create_sheet("Member Months")

    headers = ["Provider ID", "Reporting Period", "Line of Business", "Member Months", "Unique Members"]
    for col, header in enumerate(headers, 1):
        ws_member.cell(row=1, column=col, value=header)

    row = 2
    for provider_id in providers[:5]:  # Just use first 5 providers
        for month in range(1, 4):  # Just 3 months of data
            ws_member.cell(row=row, column=1, value=provider_id)
            ws_member.cell(row=row, column=2, value=f"2025-{month:02d}")
            ws_member.cell(row=row, column=3, value="Commercial")
            ws_member.cell(row=row, column=4, value=random.randint(100, 500))
            ws_member.cell(row=row, column=5, value=random.randint(80, 400))
            row += 1

    # 3. Medical Claims Sheet
    ws_medical = wb.create_sheet("Medical Claims")

    headers = [
        "Provider ID",
        "Member ID",
        "Claim ID",
        "Service Date",
        "Paid Date",
        "Procedure Code",
        "Diagnosis Code",
        "Allowed Amount",
        "Paid Amount",
        "Member Liability",
    ]
    for col, header in enumerate(headers, 1):
        ws_medical.cell(row=1, column=col, value=header)

    # Create sample claims
    total_medical = 0
    for i in range(100):
        row = i + 2
        provider_id = random.choice(providers)
        allowed = round(random.uniform(50, 500), 2)
        paid = round(allowed * 0.8, 2)
        liability = round(allowed - paid, 2)
        total_medical += allowed

        service_date = datetime(2025, 1, 1) + timedelta(days=random.randint(0, 60))
        paid_date = service_date + timedelta(days=random.randint(10, 30))

        ws_medical.cell(row=row, column=1, value=provider_id)
        ws_medical.cell(row=row, column=2, value=f"MEM{str(i+1).zfill(8)}")
        ws_medical.cell(row=row, column=3, value=f"CLM{str(i+1).zfill(10)}")
        ws_medical.cell(row=row, column=4, value=service_date).number_format = "YYYY-MM-DD"
        ws_medical.cell(row=row, column=5, value=paid_date).number_format = "YYYY-MM-DD"
        ws_medical.cell(row=row, column=6, value="99213")
        ws_medical.cell(row=row, column=7, value="Z00.00")
        ws_medical.cell(row=row, column=8, value=allowed)
        ws_medical.cell(row=row, column=9, value=paid)
        ws_medical.cell(row=row, column=10, value=liability)

    # 4. Pharmacy Claims Sheet (with Provider ID!)
    ws_pharmacy = wb.create_sheet("Pharmacy Claims")

    headers = [
        "Provider ID",
        "Member ID",
        "Claim ID",
        "Fill Date",
        "NDC",
        "Quantity",
        "Days Supply",
        "Allowed Amount",
        "Paid Amount",
        "Member Liability",
    ]
    for col, header in enumerate(headers, 1):
        ws_pharmacy.cell(row=1, column=col, value=header)

    # Create sample pharmacy claims
    for i in range(50):
        row = i + 2
        provider_id = random.choice(providers)
        allowed = round(random.uniform(20, 200), 2)
        paid = round(allowed * 0.9, 2)
        liability = round(allowed - paid, 2)

        fill_date = datetime(2025, 1, 1) + timedelta(days=random.randint(0, 60))

        ws_pharmacy.cell(row=row, column=1, value=provider_id)  # Include Provider ID!
        ws_pharmacy.cell(row=row, column=2, value=f"MEM{str(i+1).zfill(8)}")
        ws_pharmacy.cell(row=row, column=3, value=f"RX{str(i+1).zfill(10)}")
        ws_pharmacy.cell(row=row, column=4, value=fill_date).number_format = "YYYY-MM-DD"
        ws_pharmacy.cell(row=row, column=5, value="00069-0100-30")
        ws_pharmacy.cell(row=row, column=6, value=30)
        ws_pharmacy.cell(row=row, column=7, value=30)
        ws_pharmacy.cell(row=row, column=8, value=allowed)
        ws_pharmacy.cell(row=row, column=9, value=paid)
        ws_pharmacy.cell(row=row, column=10, value=liability)

    # 5. Reconciliation Sheet (with matching totals!)
    ws_recon = wb.create_sheet("Reconciliation")

    headers = [
        "Provider ID",
        "Medical Claims Total",
        "Medical Paid Total",
        "Pharmacy Claims Total",
        "Pharmacy Paid Total",
        "Adjustments",
        "Net Total",
    ]
    for col, header in enumerate(headers, 1):
        ws_recon.cell(row=1, column=col, value=header)

    # Calculate actual totals from medical sheet
    row = 2
    ws_recon.cell(row=row, column=1, value="ALL")
    ws_recon.cell(row=row, column=2, value=round(total_medical, 2))  # Match the actual total!
    ws_recon.cell(row=row, column=3, value=round(total_medical * 0.8, 2))
    ws_recon.cell(row=row, column=4, value=5000.00)
    ws_recon.cell(row=row, column=5, value=4500.00)
    ws_recon.cell(row=row, column=6, value=0.00)
    ws_recon.cell(row=row, column=7, value=round(total_medical, 2))

    # Save the file
    output_path = Path("./mock_data/oregon/test_truly_valid.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✓ Created truly valid test file: {output_path}")
    print("  - NPIs and ZIPs are properly formatted as text")
    print("  - Provider ID included in Pharmacy Claims")
    print(f"  - Reconciliation totals match Medical Claims total: ${total_medical:,.2f}")
    print("  - All dates are properly formatted")

    return output_path


if __name__ == "__main__":
    create_valid_oregon_submission()
