"""Excel error annotation - highlights validation errors directly in Excel files."""

import shutil
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .validation_results import Severity, ValidationIssue, ValidationResults


class ExcelAnnotator:
    """Annotate Excel files with validation errors."""

    # Color schemes for different severity levels
    SEVERITY_COLORS = {
        Severity.ERROR: {
            "fill": PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
            "font": Font(color="CC0000", bold=True),
            "border_color": "CC0000",
        },
        Severity.WARNING: {
            "fill": PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
            "font": Font(color="FF6600", bold=True),
            "border_color": "FF6600",
        },
        Severity.INFO: {
            "fill": PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
            "font": Font(color="0066CC", italic=True),
            "border_color": "0066CC",
        },
    }

    def __init__(self):
        self.workbook = None
        self.annotations_count = 0

    def annotate_file(
        self, source_file: str, validation_results: ValidationResults, output_file: Optional[str] = None
    ) -> Path:
        """
        Annotate an Excel file with validation results.

        Args:
            source_file: Path to the original Excel file
            validation_results: Validation results to annotate
            output_file: Path for annotated file (if None, adds '_annotated' suffix)

        Returns:
            Path to the annotated Excel file
        """
        source_path = Path(source_file)

        # Determine output path
        if output_file:
            output_path = Path(output_file)
        else:
            output_path = source_path.parent / f"{source_path.stem}_annotated{source_path.suffix}"

        # Copy source file to output
        shutil.copy2(source_path, output_path)

        # Load workbook
        self.workbook = load_workbook(output_path)
        self.annotations_count = 0

        # Add summary sheet
        self._add_summary_sheet(validation_results)

        # Process all issues
        all_issues = (
            [(issue, Severity.ERROR) for issue in validation_results.errors]
            + [(issue, Severity.WARNING) for issue in validation_results.warnings]
            + [(issue, Severity.INFO) for issue in validation_results.info]
        )

        # Group issues by location
        issues_by_location = self._group_issues_by_location(all_issues)

        # Annotate each location
        for location, issues in issues_by_location.items():
            self._annotate_location(location, issues)

        # Save annotated workbook
        self.workbook.save(output_path)

        return output_path

    def _add_summary_sheet(self, results: ValidationResults):
        """Add a summary sheet with validation results overview."""
        # Create or get summary sheet
        if "Validation Summary" in self.workbook.sheetnames:
            summary_sheet = self.workbook["Validation Summary"]
            summary_sheet.delete_rows(1, summary_sheet.max_row)
        else:
            summary_sheet = self.workbook.create_sheet("Validation Summary", 0)

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        # Add title
        summary_sheet.merge_cells("A1:E1")
        title_cell = summary_sheet["A1"]
        title_cell.value = "CGT Validation Summary"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")

        # Add metadata
        summary_sheet["A3"] = "State:"
        summary_sheet["B3"] = results.state.title()
        summary_sheet["A4"] = "Year:"
        summary_sheet["B4"] = results.year
        summary_sheet["A5"] = "Status:"
        summary_sheet["B5"] = "PASSED" if results.is_valid() else "FAILED"

        # Style status cell
        status_cell = summary_sheet["B5"]
        if results.is_valid():
            status_cell.font = Font(color="008000", bold=True)
        else:
            status_cell.font = Font(color="CC0000", bold=True)

        # Add summary counts
        summary_sheet["A7"] = "Issue Summary"
        summary_sheet["A7"].font = Font(bold=True, size=12)

        summary_sheet["A8"] = "Errors:"
        summary_sheet["B8"] = len(results.errors)
        summary_sheet["B8"].font = Font(color="CC0000", bold=True)

        summary_sheet["A9"] = "Warnings:"
        summary_sheet["B9"] = len(results.warnings)
        summary_sheet["B9"].font = Font(color="FF6600", bold=True)

        summary_sheet["A10"] = "Info:"
        summary_sheet["B10"] = len(results.info)
        summary_sheet["B10"].font = Font(color="0066CC", bold=True)

        # Add issues table
        if results.errors or results.warnings:
            # Table headers
            headers = ["Severity", "Location", "Code", "Message", "Sheet Link"]
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=13, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            # Add issues
            row = 14
            for error_issue in results.errors:
                self._add_issue_row(summary_sheet, row, error_issue, Severity.ERROR)
                row += 1

            for warning_issue in results.warnings:
                self._add_issue_row(summary_sheet, row, warning_issue, Severity.WARNING)
                row += 1

        # Adjust column widths
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            if col == 4:  # Message column
                summary_sheet.column_dimensions[column_letter].width = 60
            else:
                summary_sheet.column_dimensions[column_letter].width = 20

    def _add_issue_row(self, sheet, row: int, issue: ValidationIssue, severity: Severity):
        """Add an issue row to the summary sheet."""
        # Severity
        sheet.cell(row=row, column=1).value = severity.value.upper()
        sheet.cell(row=row, column=1).font = self.SEVERITY_COLORS[severity]["font"]

        # Location
        sheet.cell(row=row, column=2).value = issue.location

        # Code
        sheet.cell(row=row, column=3).value = issue.code

        # Message
        sheet.cell(row=row, column=4).value = issue.message

        # Sheet link (if applicable)
        location_parts = issue.location.split(".")
        if location_parts[0] in self.workbook.sheetnames:
            link_cell = sheet.cell(row=row, column=5)
            link_cell.value = f"Go to {location_parts[0]}"
            link_cell.font = Font(color="0000FF", underline="single")
            # Note: Actual hyperlink would require more complex implementation

    def _group_issues_by_location(
        self, issues: Sequence[Tuple[ValidationIssue, Severity]]
    ) -> Dict[str, List[Tuple[ValidationIssue, Severity]]]:
        """Group issues by their location."""
        grouped: Dict[str, List[Tuple[ValidationIssue, Severity]]] = {}

        for issue, severity in issues:
            if issue.location not in grouped:
                grouped[issue.location] = []
            grouped[issue.location].append((issue, severity))

        return grouped

    def _annotate_location(self, location: str, issues: List[Tuple[ValidationIssue, Severity]]):
        """Annotate a specific location with issues."""
        # Parse location (e.g., "Sheet1.Column1" or "Sheet1.A5")
        parts = location.split(".")

        if len(parts) == 0:
            return

        sheet_name = parts[0]

        # Handle different location types
        if sheet_name == "file":
            # File-level issue - add to summary only
            return
        elif sheet_name == "sheets":
            # Sheet structure issue - add to summary only
            return
        elif sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]

            if len(parts) == 1:
                # Sheet-level annotation
                self._annotate_sheet(worksheet, issues)
            elif len(parts) == 2:
                # Column or cell annotation
                import re

                # Check if it's a valid Excel cell reference (e.g., A1, B2, AA10, etc.)
                cell_ref_pattern = r"^[A-Z]+[0-9]+$"
                if re.match(cell_ref_pattern, parts[1].upper()):  # Cell reference
                    self._annotate_cell(worksheet, parts[1], issues)
                else:
                    # Column annotation
                    self._annotate_column(worksheet, parts[1], issues)

    def _annotate_sheet(self, worksheet, issues: List[Tuple[ValidationIssue, Severity]]):
        """Add sheet-level annotations."""
        # Add a comment to cell A1 with all issues
        cell = worksheet["A1"]

        # Combine all issue messages
        messages = []
        highest_severity = Severity.INFO

        for issue, severity in issues:
            messages.append(f"[{issue.code}] {issue.message}")
            if severity == Severity.ERROR:
                highest_severity = Severity.ERROR
            elif severity == Severity.WARNING and highest_severity != Severity.ERROR:
                highest_severity = Severity.WARNING

        # Add comment
        comment_text = "Validation Issues:\n" + "\n".join(messages)
        cell.comment = Comment(comment_text, "CGT Validator")

        # Apply styling to indicate issues
        style = self.SEVERITY_COLORS[highest_severity]
        cell.fill = style["fill"]

        self.annotations_count += 1

    def _annotate_column(self, worksheet, column_name: str, issues: List[Tuple[ValidationIssue, Severity]]):
        """Annotate all cells in a column that have issues."""
        # Find column index
        header_row = 1
        column_index = None

        for col in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=header_row, column=col).value == column_name:
                column_index = col
                break

        if not column_index:
            return

        # Apply styling to column header
        header_cell = worksheet.cell(row=header_row, column=column_index)

        # Determine highest severity
        highest_severity = Severity.INFO
        for issue, severity in issues:
            if severity == Severity.ERROR:
                highest_severity = Severity.ERROR
            elif severity == Severity.WARNING and highest_severity != Severity.ERROR:
                highest_severity = Severity.WARNING

        # Apply style
        style = self.SEVERITY_COLORS[highest_severity]
        header_cell.fill = style["fill"]
        header_cell.font = style["font"]

        # Add border
        border = Border(
            left=Side(style="thin", color=style["border_color"]),
            right=Side(style="thin", color=style["border_color"]),
            top=Side(style="thin", color=style["border_color"]),
            bottom=Side(style="thin", color=style["border_color"]),
        )
        header_cell.border = border

        # Add comment with all issues
        messages = [f"[{issue.code}] {issue.message}" for issue, _ in issues]
        comment_text = "Validation Issues:\n" + "\n".join(messages)
        header_cell.comment = Comment(comment_text, "CGT Validator")

        # Also highlight cells mentioned in error messages
        for issue, severity in issues:
            # Look for row numbers in message
            import re

            row_matches = re.findall(r"rows?:?\s*\[?(\d+(?:,\s*\d+)*)", issue.message, re.IGNORECASE)
            if row_matches:
                for match in row_matches:
                    rows = [int(r.strip()) for r in match.split(",")]
                    for row_num in rows[:10]:  # Limit to first 10 rows
                        if row_num <= worksheet.max_row:
                            cell = worksheet.cell(row=row_num + 1, column=column_index)  # +1 for header
                            cell.fill = style["fill"]
                            if not cell.comment:
                                cell.comment = Comment(f"[{issue.code}] Issue detected", "CGT Validator")

        self.annotations_count += len(issues)

    def _annotate_cell(self, worksheet, cell_ref: str, issues: List[Tuple[ValidationIssue, Severity]]):
        """Annotate a specific cell."""
        try:
            cell = worksheet[cell_ref]
        except (KeyError, ValueError):
            return

        # Determine highest severity
        highest_severity = Severity.INFO
        messages = []

        for issue, severity in issues:
            messages.append(f"[{issue.code}] {issue.message}")
            if severity == Severity.ERROR:
                highest_severity = Severity.ERROR
            elif severity == Severity.WARNING and highest_severity != Severity.ERROR:
                highest_severity = Severity.WARNING

        # Apply style
        style = self.SEVERITY_COLORS[highest_severity]
        cell.fill = style["fill"]
        cell.font = style["font"]

        # Add border
        border = Border(
            left=Side(style="medium", color=style["border_color"]),
            right=Side(style="medium", color=style["border_color"]),
            top=Side(style="medium", color=style["border_color"]),
            bottom=Side(style="medium", color=style["border_color"]),
        )
        cell.border = border

        # Add comment
        comment_text = "Validation Issues:\n" + "\n".join(messages)
        cell.comment = Comment(comment_text, "CGT Validator")

        self.annotations_count += 1

    def get_annotations_count(self) -> int:
        """Get the number of annotations made."""
        return int(self.annotations_count)
