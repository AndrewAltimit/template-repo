"""Base validator class for all state-specific validators."""

from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from reporters.validation_results import Severity, ValidationResults


class ValidatorBase(ABC):
    """Abstract base class for state-specific validators."""

    def __init__(self, state: str, year: Optional[int] = None):
        self.state = state
        self.year = year or datetime.now().year
        self.requirements = self._load_requirements()

    @abstractmethod
    def _load_requirements(self) -> Dict[str, Any]:
        """Load state-specific requirements."""

    def validate_file(self, filepath: str) -> ValidationResults:
        """Main validation entry point for end users."""
        results = ValidationResults(state=self.state, year=self.year)
        file_path = Path(filepath)

        # Check file exists and is readable
        if not self._check_file_access(file_path, results):
            return results

        try:
            # Load Excel file
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            excel_data = pd.ExcelFile(file_path)

            # Run all validations
            self._validate_structure(workbook, excel_data, results)
            self._validate_sheet_names(workbook, results)
            self._validate_mandatory_fields(excel_data, results)
            self._validate_data_types(excel_data, results)
            self._validate_business_rules(excel_data, results)
            self._validate_cross_references(excel_data, results)

            # State-specific validations
            self._run_state_specific_validations(excel_data, results)

        except Exception as e:  # pylint: disable=broad-exception-caught
            results.add_error(
                code="FILE_READ_ERROR",
                message=f"Error reading file: {str(e)}",
                location="file",
                severity=Severity.ERROR,
            )

        return results

    def _check_file_access(self, file_path: Path, results: ValidationResults) -> bool:
        """Check if file exists and is accessible."""
        if not file_path.exists():
            results.add_error(
                code="FILE_NOT_FOUND", message=f"File not found: {file_path}", location="file", severity=Severity.ERROR
            )
            return False

        if not file_path.suffix.lower() in [".xlsx", ".xlsm"]:
            results.add_error(
                code="INVALID_FILE_TYPE",
                message=f"Invalid file type. Expected .xlsx or .xlsm, got {file_path.suffix}",
                location="file",
                severity=Severity.ERROR,
            )
            return False

        if not file_path.stat().st_size > 0:
            results.add_error(code="EMPTY_FILE", message="File is empty", location="file", severity=Severity.ERROR)
            return False

        return True

    def _validate_structure(self, workbook, excel_data, results: ValidationResults):  # pylint: disable=unused-argument
        """Validate overall file structure."""
        # Check for minimum number of sheets
        min_sheets = self.requirements.get("min_sheets", 1)
        if len(workbook.sheetnames) < min_sheets:
            results.add_error(
                code="INSUFFICIENT_SHEETS",
                message=f"File must have at least {min_sheets} sheets, found {len(workbook.sheetnames)}",
                location="file",
                severity=Severity.ERROR,
            )

    def _validate_sheet_names(self, workbook, results: ValidationResults):
        """Validate that required sheets are present."""
        required_sheets = self.requirements.get("required_sheets", [])
        present_sheets = set(workbook.sheetnames)

        for required_sheet in required_sheets:
            if required_sheet not in present_sheets:
                results.add_error(
                    code="MISSING_SHEET",
                    message=f"Required sheet '{required_sheet}' not found",
                    location="sheets",
                    severity=Severity.ERROR,
                )

        # Check for unexpected sheets
        expected_sheets = set(self.requirements.get("all_sheets", required_sheets))
        unexpected_sheets = present_sheets - expected_sheets

        for sheet in unexpected_sheets:
            results.add_warning(
                code="UNEXPECTED_SHEET",
                message=f"Unexpected sheet '{sheet}' found",
                location="sheets",
                severity=Severity.WARNING,
            )

    def _validate_mandatory_fields(self, excel_data: pd.ExcelFile, results: ValidationResults):
        """Validate that mandatory fields are present and not empty."""
        mandatory_fields = self.requirements.get("mandatory_fields", {})

        for sheet_name, fields in mandatory_fields.items():
            if sheet_name not in excel_data.sheet_names:
                continue

            df = excel_data.parse(sheet_name)

            # Check column presence
            for field in fields:
                if field not in df.columns:
                    results.add_error(
                        code="MISSING_COLUMN",
                        message=f"Required column '{field}' not found",
                        location=f"{sheet_name}",
                        severity=Severity.ERROR,
                    )
                else:
                    # Check for empty values
                    empty_rows = df[df[field].isna() | (df[field] == "")].index.tolist()
                    if empty_rows:
                        results.add_error(
                            code="EMPTY_MANDATORY_FIELD",
                            message=(
                                f"Column '{field}' has empty values in rows: "
                                f"{empty_rows[:5]}{'...' if len(empty_rows) > 5 else ''}"
                            ),
                            location=f"{sheet_name}.{field}",
                            severity=Severity.ERROR,
                        )

    def _validate_data_types(self, excel_data: pd.ExcelFile, results: ValidationResults):
        """Validate data types for each column."""
        data_types = self.requirements.get("data_types", {})

        for sheet_name, column_types in data_types.items():
            if sheet_name not in excel_data.sheet_names:
                continue

            df = excel_data.parse(sheet_name)

            for column, expected_type in column_types.items():
                if column not in df.columns:
                    continue

                # Check data types
                invalid_rows = []
                for idx, value in df[column].items():
                    if pd.isna(value):
                        continue

                    if not self._check_data_type(value, expected_type):
                        invalid_rows.append(idx)

                if invalid_rows:
                    results.add_error(
                        code="INVALID_DATA_TYPE",
                        message=(
                            f"Column '{column}' expects {expected_type}, found invalid values in rows: "
                            f"{invalid_rows[:5]}{'...' if len(invalid_rows) > 5 else ''}"
                        ),
                        location=f"{sheet_name}.{column}",
                        severity=Severity.ERROR,
                    )

    def _check_data_type(self, value: Any, expected_type: str) -> bool:
        """Check if a value matches the expected data type."""
        if expected_type == "numeric":
            try:
                float(value)
                return True
            except (ValueError, TypeError):
                return False
        elif expected_type == "date":
            return isinstance(value, (datetime, pd.Timestamp))
        elif expected_type == "text":
            # Accept strings or numbers that can be converted to strings
            # This handles cases where Excel reads numeric-looking strings as numbers
            return isinstance(value, (str, int, float))
        elif expected_type == "integer":
            try:
                return float(value).is_integer()
            except (ValueError, TypeError):
                return False
        else:
            return True

    @abstractmethod
    def _validate_business_rules(self, excel_data: pd.ExcelFile, results: ValidationResults):
        """Validate state-specific business rules."""

    @abstractmethod
    def _validate_cross_references(self, excel_data: pd.ExcelFile, results: ValidationResults):
        """Validate cross-references between sheets."""

    @abstractmethod
    def _run_state_specific_validations(self, excel_data: pd.ExcelFile, results: ValidationResults):
        """Run additional state-specific validations."""

    def validate_numeric_range(
        self,
        df: pd.DataFrame,
        column: str,
        min_val: Optional[float],
        max_val: Optional[float],
        sheet_name: str,
        results: ValidationResults,
    ):
        """Helper to validate numeric ranges."""
        if column not in df.columns:
            return

        if min_val is not None:
            invalid = df[df[column] < min_val]
            if not invalid.empty:
                results.add_error(
                    code="VALUE_TOO_LOW",
                    message=f"Column '{column}' has values below minimum {min_val} in rows: {invalid.index.tolist()[:5]}",
                    location=f"{sheet_name}.{column}",
                    severity=Severity.ERROR,
                )

        if max_val is not None:
            invalid = df[df[column] > max_val]
            if not invalid.empty:
                results.add_error(
                    code="VALUE_TOO_HIGH",
                    message=f"Column '{column}' has values above maximum {max_val} in rows: {invalid.index.tolist()[:5]}",
                    location=f"{sheet_name}.{column}",
                    severity=Severity.ERROR,
                )

    def validate_allowed_values(
        self, df: pd.DataFrame, column: str, allowed_values: List[Any], sheet_name: str, results: ValidationResults
    ):
        """Helper to validate against allowed values."""
        if column not in df.columns:
            return

        invalid = df[~df[column].isin(allowed_values)]
        if not invalid.empty:
            results.add_error(
                code="INVALID_VALUE",
                message=(
                    f"Column '{column}' has invalid values. Allowed: {allowed_values}. "
                    f"Found in rows: {invalid.index.tolist()[:5]}"
                ),
                location=f"{sheet_name}.{column}",
                severity=Severity.ERROR,
            )

    def validate_unique_values(self, df: pd.DataFrame, column: str, sheet_name: str, results: ValidationResults):
        """Helper to validate uniqueness."""
        if column not in df.columns:
            return

        duplicates = df[df.duplicated(subset=[column], keep=False)]
        if not duplicates.empty:
            results.add_error(
                code="DUPLICATE_VALUES",
                message=f"Column '{column}' has duplicate values in rows: {duplicates.index.tolist()[:10]}",
                location=f"{sheet_name}.{column}",
                severity=Severity.ERROR,
            )
