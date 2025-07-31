"""Parser for extracting requirements from CGT documents (PDFs and Excel templates)."""

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import pdfplumber
from openpyxl import load_workbook


class RequirementsParser:
    """Extract validation requirements from CGT documents."""

    def __init__(self, state: str, output_dir: Optional[Path] = None):
        self.state = state
        self.output_dir = output_dir or Path(f"./requirements/{state}")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def parse_document(self, file_path: Path) -> Dict[str, Any]:
        """Parse a document and extract requirements."""
        if file_path.suffix.lower() == ".pdf":
            return self.parse_pdf(file_path)
        elif file_path.suffix.lower() in [".xlsx", ".xlsm"]:
            return self.parse_excel_template(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_path.suffix}")

    def parse_pdf(self, pdf_path: Path) -> Dict[str, Any]:
        """Extract requirements from PDF specification manuals."""
        requirements = {
            "source_file": str(pdf_path),
            "file_type": "pdf",
            "sheets": {},
            "fields": {},
            "validation_rules": [],
            "business_rules": [],
            "data_dictionary": {},
            "cross_references": [],
        }

        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            tables = []

            for page_num, page in enumerate(pdf.pages):
                # Extract text
                page_text = page.extract_text() or ""
                full_text += page_text + "\n"

                # Extract tables
                page_tables = page.extract_tables()
                if page_tables:
                    for table in page_tables:
                        tables.append({"page": page_num + 1, "data": table})

            # Parse sheet information
            requirements["sheets"] = self._extract_sheet_info(full_text)

            # Parse field definitions from tables
            requirements["fields"] = self._extract_field_definitions(tables, full_text)

            # Parse validation rules
            requirements["validation_rules"] = self._extract_validation_rules(full_text)

            # Parse business rules
            requirements["business_rules"] = self._extract_business_rules(full_text)

            # Extract data dictionary
            requirements["data_dictionary"] = self._extract_data_dictionary(tables, full_text)

            # Extract cross-reference requirements
            requirements["cross_references"] = self._extract_cross_references(full_text)

        return requirements

    def parse_excel_template(self, excel_path: Path) -> Dict[str, Any]:
        """Extract requirements from Excel template files."""
        requirements = {
            "source_file": str(excel_path),
            "file_type": "excel",
            "sheets": {},
            "fields": {},
            "validation_rules": [],
            "data_validations": {},
            "formulas": {},
            "named_ranges": {},
        }

        # Load workbook
        wb = load_workbook(excel_path, data_only=False)

        # Extract sheet information
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "columns": [],
                "row_count": ws.max_row,
                "col_count": ws.max_column,
                "has_data_validation": False,
                "has_formulas": False,
            }

            # Get column headers (assume first row)
            if ws.max_row > 0:
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                sheet_info["columns"] = headers

            # Check for data validations
            if ws.data_validations:
                sheet_info["has_data_validation"] = True
                requirements["data_validations"][sheet_name] = self._extract_excel_validations(ws)

            # Check for formulas
            formula_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells.append({"cell": cell.coordinate, "formula": cell.value})

            if formula_cells:
                sheet_info["has_formulas"] = True
                requirements["formulas"][sheet_name] = formula_cells

            requirements["sheets"][sheet_name] = sheet_info

        # Extract named ranges
        if wb.defined_names:
            for defn in wb.defined_names.definedName:
                requirements["named_ranges"][defn.name] = {"range": defn.value, "scope": defn.localSheetId}

        # Extract field requirements from data
        requirements["fields"] = self._extract_fields_from_excel(wb)

        # Generate validation rules based on template structure
        requirements["validation_rules"] = self._generate_validation_rules_from_excel(requirements)

        return requirements

    def _extract_sheet_info(self, text: str) -> Dict[str, Dict]:
        """Extract sheet/tab information from text."""
        sheets = {}

        # Common patterns for sheet descriptions
        sheet_patterns = [
            r"(?:Sheet|Tab|Worksheet)[\s:]+([A-Za-z0-9\s]+)(?:\n|:)",
            r"(?:The\s+)?([A-Za-z0-9\s]+)\s+(?:sheet|tab|worksheet)",
            r"\"([A-Za-z0-9\s]+)\"\s+(?:sheet|tab)",
        ]

        for pattern in sheet_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                sheet_name = match.group(1).strip()
                if len(sheet_name) > 2 and len(sheet_name) < 50:  # Reasonable sheet name
                    sheets[sheet_name] = {
                        "required": "required" in text[max(0, match.start() - 50) : match.end() + 50].lower(),
                        "description": self._extract_description(text, match.start()),
                    }

        return sheets

    def _extract_field_definitions(self, tables: List[Dict], text: str) -> Dict[str, List[Dict]]:
        """Extract field definitions from tables and text."""
        fields = {}

        # Look for field definition tables
        for table_info in tables:
            table = table_info["data"]
            if not table or len(table) < 2:
                continue

            # Check if this looks like a field definition table
            headers = table[0] if table else []
            if any(
                h and any(keyword in str(h).lower() for keyword in ["field", "column", "attribute", "name"])
                for h in headers
            ):
                # Parse field definitions
                field_list = []
                for row in table[1:]:
                    if len(row) >= 2 and row[0]:  # At least field name and some info
                        field_def = {
                            "name": str(row[0]).strip(),
                            "type": self._extract_field_type(row),
                            "required": self._is_field_required(row),
                            "description": self._extract_field_description(row),
                            "format": self._extract_field_format(row),
                            "allowed_values": self._extract_allowed_values(row),
                        }
                        field_list.append(field_def)

                if field_list:
                    # Try to determine which sheet these fields belong to
                    sheet_name = self._determine_sheet_for_fields(table_info["page"], text)
                    if sheet_name not in fields:
                        fields[sheet_name] = []
                    fields[sheet_name].extend(field_list)

        return fields

    def _extract_validation_rules(self, text: str) -> List[Dict[str, str]]:
        """Extract validation rules from text."""
        rules = []

        # Patterns for validation rules
        rule_patterns = [
            r"(?:must|should|shall)\s+(?:be|have|contain)\s+([^.]+)",
            r"(?:valid|invalid)\s+(?:values?|formats?)\s+(?:are|is|include)\s*:?\s*([^.]+)",
            r"(?:format|pattern):\s*([^.\n]+)",
            r"(?:minimum|maximum|min|max)\s+(?:value|length|size)\s*:?\s*([0-9]+)",
            r"(?:required|mandatory|optional)\s+(?:field|column|attribute)s?\s*:?\s*([^.\n]+)",
        ]

        for pattern in rule_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                rule_text = match.group(0).strip()
                rules.append(
                    {
                        "type": self._classify_rule(rule_text),
                        "text": rule_text,
                        "context": text[max(0, match.start() - 100) : match.end() + 100],
                    }
                )

        return rules

    def _extract_business_rules(self, text: str) -> List[Dict[str, Any]]:
        """Extract business rules from text."""
        rules = []

        # Look for numbered rules or business logic sections
        business_patterns = [
            r"(?:Business Rule|BR)\s*#?(\d+)\s*:?\s*([^.\n]+)",
            r"(\d+)\.\s*(?:The|A|An)?\s*([A-Z][^.]+(?:must|should|shall)[^.]+)",
            r"(?:Rule|Requirement)\s*:?\s*([^.\n]+(?:must|should|shall)[^.]+)",
        ]

        for pattern in business_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                rule_text = match.group(0).strip()
                rules.append(
                    {
                        "id": f"BR_{len(rules)+1}",
                        "text": rule_text,
                        "type": "business_rule",
                        "severity": "error" if "must" in rule_text.lower() else "warning",
                    }
                )

        return rules

    def _extract_data_dictionary(self, tables: List[Dict], text: str) -> Dict[str, Dict]:
        """Extract data dictionary information."""
        data_dict = {}

        # Look for data dictionary tables
        for table_info in tables:
            table = table_info["data"]
            if not table:
                continue

            # Check if this is a data dictionary table
            if self._is_data_dictionary_table(table):
                for row in table[1:]:  # Skip header
                    if len(row) >= 2 and row[0]:
                        field_name = str(row[0]).strip()
                        data_dict[field_name] = {
                            "type": self._extract_data_type(row),
                            "length": self._extract_field_length(row),
                            "format": self._extract_field_format(row),
                            "description": self._extract_field_description(row),
                            "allowed_values": self._extract_allowed_values(row),
                        }

        return data_dict

    def _extract_cross_references(self, text: str) -> List[Dict[str, str]]:
        """Extract cross-reference requirements."""
        cross_refs = []

        # Patterns for cross-references
        ref_patterns = [
            r"(?:must|should)\s+(?:match|reference|correspond to)\s+(?:the\s+)?([^.]+)",
            r"(?:foreign key|reference|link)\s+to\s+([^.]+)",
            r"([A-Za-z\s]+)\s+(?:from|in)\s+([A-Za-z\s]+)\s+(?:sheet|tab)",
        ]

        for pattern in ref_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                cross_refs.append(
                    {
                        "type": "cross_reference",
                        "description": match.group(0).strip(),
                        "source": self._extract_source_field(match.group(0)),
                        "target": self._extract_target_field(match.group(0)),
                    }
                )

        return cross_refs

    def _extract_excel_validations(self, worksheet) -> List[Dict]:
        """Extract data validation rules from Excel worksheet."""
        validations = []

        for dv in worksheet.data_validations.dataValidation:
            validation = {
                "type": dv.type,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allowBlank,
                "ranges": str(dv.ranges),
                "error_message": dv.error,
                "error_title": dv.errorTitle,
            }

            # Extract validation specifics
            if dv.type == "list":
                validation["allowed_values"] = self._parse_list_validation(dv.formula1)
            elif dv.type in ["whole", "decimal"]:
                validation["min_value"] = dv.formula1
                validation["max_value"] = dv.formula2

            validations.append(validation)

        return validations

    def _extract_fields_from_excel(self, workbook) -> Dict[str, List[Dict]]:
        """Extract field information from Excel workbook."""
        fields = {}

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            if ws.max_row == 0:
                continue

            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(
                        {
                            "name": str(header).strip(),
                            "column": col,
                            "data_type": self._infer_data_type_from_column(ws, col),
                            "has_validation": self._check_column_validation(ws, col),
                            "sample_values": self._get_sample_values(ws, col),
                        }
                    )

            if headers:
                fields[sheet_name] = headers

        return fields

    def _generate_validation_rules_from_excel(self, requirements: Dict) -> List[Dict]:
        """Generate validation rules based on Excel template structure."""
        rules = []

        # Required sheets rule
        sheet_names = list(requirements["sheets"].keys())
        if sheet_names:
            rules.append(
                {
                    "type": "structure",
                    "rule": "required_sheets",
                    "sheets": sheet_names,
                    "description": f"File must contain these sheets: {', '.join(sheet_names)}",
                }
            )

        # Field rules
        for sheet_name, fields in requirements["fields"].items():
            for field in fields:
                # Required field rule
                rules.append(
                    {
                        "type": "field",
                        "rule": "required_field",
                        "sheet": sheet_name,
                        "field": field["name"],
                        "description": f"Sheet '{sheet_name}' must contain field '{field['name']}'",
                    }
                )

                # Data type rule
                if field.get("data_type"):
                    rules.append(
                        {
                            "type": "data_type",
                            "rule": "field_type",
                            "sheet": sheet_name,
                            "field": field["name"],
                            "expected_type": field["data_type"],
                            "description": f"Field '{field['name']}' in '{sheet_name}' must be {field['data_type']}",
                        }
                    )

        # Data validation rules
        for sheet_name, validations in requirements.get("data_validations", {}).items():
            for validation in validations:
                rules.append(
                    {
                        "type": "validation",
                        "rule": validation["type"],
                        "sheet": sheet_name,
                        "ranges": validation["ranges"],
                        "details": validation,
                        "description": f"Validation rule in '{sheet_name}': {validation}",
                    }
                )

        return rules

    def save_requirements(self, requirements: Dict[str, Any], filename: Optional[str] = None):
        """Save parsed requirements to JSON file."""
        if not filename:
            filename = f"{self.state}_requirements_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.json"

        output_path = self.output_dir / filename
        with open(output_path, "w") as f:
            json.dump(requirements, f, indent=2, default=str)

        print(f"Requirements saved to: {output_path}")
        return output_path

    # Helper methods
    def _extract_description(self, text: str, position: int) -> str:
        """Extract description text near a position."""
        start = max(0, position - 200)
        end = min(len(text), position + 200)
        context = text[start:end]

        # Clean up
        context = re.sub(r"\s+", " ", context).strip()
        return context[:200] + "..." if len(context) > 200 else context

    def _extract_field_type(self, row: List) -> str:
        """Extract field type from a table row."""
        type_keywords = {
            "text": ["text", "string", "char", "varchar"],
            "numeric": ["numeric", "number", "decimal", "float", "double"],
            "integer": ["integer", "int", "whole"],
            "date": ["date", "datetime", "timestamp"],
            "boolean": ["boolean", "bool", "yes/no", "true/false"],
        }

        row_text = " ".join(str(cell).lower() for cell in row if cell)

        for data_type, keywords in type_keywords.items():
            if any(keyword in row_text for keyword in keywords):
                return data_type

        return "text"  # Default

    def _is_field_required(self, row: List) -> bool:
        """Check if field is marked as required."""
        row_text = " ".join(str(cell).lower() for cell in row if cell)
        return any(keyword in row_text for keyword in ["required", "mandatory", "must", "yes"])

    def _extract_field_description(self, row: List) -> str:
        """Extract field description from row."""
        # Usually in later columns
        if len(row) >= 3:
            for i in range(2, len(row)):
                if row[i] and len(str(row[i])) > 10:
                    return str(row[i]).strip()
        return ""

    def _extract_field_format(self, row: List) -> str:
        """Extract format information from row."""
        row_text = " ".join(str(cell) for cell in row if cell)

        # Look for format patterns
        format_match = re.search(r"(?:format|pattern)[:\s]+([^\s,]+)", row_text, re.IGNORECASE)
        if format_match:
            return format_match.group(1)

        # Look for common formats
        if re.search(r"\d{4}-\d{2}-\d{2}", row_text):
            return "YYYY-MM-DD"
        elif re.search(r"\d{2}/\d{2}/\d{4}", row_text):
            return "MM/DD/YYYY"
        elif re.search(r"\d{10}", row_text):
            return "##########"  # 10 digits

        return ""

    def _extract_allowed_values(self, row: List) -> List[str]:
        """Extract allowed values from row."""
        allowed_values = []
        row_text = " ".join(str(cell) for cell in row if cell)

        # Look for enumerated values
        enum_match = re.search(r"(?:values|options|choices)[:\s]+([^.]+)", row_text, re.IGNORECASE)
        if enum_match:
            values_text = enum_match.group(1)
            # Parse comma or semicolon separated values
            values = re.split(r"[,;]", values_text)
            allowed_values = [v.strip() for v in values if v.strip()]

        return allowed_values

    def _determine_sheet_for_fields(self, page_num: int, text: str) -> str:
        """Try to determine which sheet fields belong to based on context."""
        # Simple heuristic - look for sheet name mentions near the page
        # In real implementation, this would be more sophisticated
        return "Unknown"

    def _classify_rule(self, rule_text: str) -> str:
        """Classify the type of validation rule."""
        rule_lower = rule_text.lower()

        if any(keyword in rule_lower for keyword in ["format", "pattern"]):
            return "format"
        elif any(keyword in rule_lower for keyword in ["required", "mandatory"]):
            return "required"
        elif any(keyword in rule_lower for keyword in ["min", "max", "length", "size"]):
            return "constraint"
        elif any(keyword in rule_lower for keyword in ["valid", "allowed", "values"]):
            return "allowed_values"
        else:
            return "business"

    def _is_data_dictionary_table(self, table: List[List]) -> bool:
        """Check if a table appears to be a data dictionary."""
        if not table or len(table) < 2:
            return False

        headers = table[0]
        header_text = " ".join(str(h).lower() for h in headers if h)

        # Check for data dictionary indicators
        dd_keywords = ["field", "column", "attribute", "type", "length", "format", "description"]
        matches = sum(1 for keyword in dd_keywords if keyword in header_text)

        return matches >= 3

    def _extract_data_type(self, row: List) -> str:
        """Extract data type from data dictionary row."""
        # Look for type column
        for cell in row:
            if cell:
                cell_lower = str(cell).lower()
                if any(t in cell_lower for t in ["varchar", "char", "text", "string"]):
                    return "text"
                elif any(t in cell_lower for t in ["int", "numeric", "decimal", "number"]):
                    return "numeric"
                elif any(t in cell_lower for t in ["date", "datetime", "timestamp"]):
                    return "date"
                elif any(t in cell_lower for t in ["bool", "boolean", "bit"]):
                    return "boolean"

        return "text"

    def _extract_field_length(self, row: List) -> Optional[int]:
        """Extract field length from row."""
        row_text = " ".join(str(cell) for cell in row if cell)

        # Look for length patterns
        length_match = re.search(r"(?:length|size|max)[:\s]*(\d+)", row_text, re.IGNORECASE)
        if length_match:
            return int(length_match.group(1))

        # Look for varchar(n) pattern
        varchar_match = re.search(r"varchar\s*\(\s*(\d+)\s*\)", row_text, re.IGNORECASE)
        if varchar_match:
            return int(varchar_match.group(1))

        return None

    def _extract_source_field(self, text: str) -> str:
        """Extract source field from cross-reference text."""
        # Simple extraction - would be more sophisticated in production
        words = text.split()
        for i, word in enumerate(words):
            if word.lower() in ["from", "in"] and i > 0:
                return words[i - 1]
        return ""

    def _extract_target_field(self, text: str) -> str:
        """Extract target field from cross-reference text."""
        # Simple extraction - would be more sophisticated in production
        words = text.split()
        for i, word in enumerate(words):
            if word.lower() in ["to", "references", "matches"] and i < len(words) - 1:
                return words[i + 1]
        return ""

    def _parse_list_validation(self, formula: str) -> List[str]:
        """Parse list validation formula to extract allowed values."""
        if not formula:
            return []

        # Remove quotes and split by comma
        formula = formula.strip("\"'")
        values = [v.strip() for v in formula.split(",")]
        return values

    def _infer_data_type_from_column(self, worksheet, col: int) -> str:
        """Infer data type from column values."""
        sample_size = min(100, worksheet.max_row - 1)

        types_found = set()
        for row in range(2, min(sample_size + 2, worksheet.max_row + 1)):
            value = worksheet.cell(row=row, column=col).value
            if value is not None:
                if isinstance(value, (int, float)):
                    types_found.add("numeric")
                elif isinstance(value, pd.Timestamp):
                    types_found.add("date")
                elif isinstance(value, bool):
                    types_found.add("boolean")
                else:
                    types_found.add("text")

        # Return most specific type
        if len(types_found) == 1:
            return list(types_found)[0]
        elif "date" in types_found:
            return "date"
        elif "numeric" in types_found:
            return "numeric"
        else:
            return "text"

    def _check_column_validation(self, worksheet, col: int) -> bool:
        """Check if column has data validation."""
        for dv in worksheet.data_validations.dataValidation:
            for range_str in str(dv.ranges).split():
                if f"${col}:" in range_str or f"${col}$" in range_str:
                    return True
        return False

    def _get_sample_values(self, worksheet, col: int, max_samples: int = 5) -> List[Any]:
        """Get sample values from a column."""
        samples = []
        unique_values = set()

        for row in range(2, min(worksheet.max_row + 1, 102)):  # Check up to 100 rows
            value = worksheet.cell(row=row, column=col).value
            if value is not None and value not in unique_values:
                unique_values.add(value)
                samples.append(value)
                if len(samples) >= max_samples:
                    break

        return samples
