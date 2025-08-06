"""Generate mock CGT data for Oregon testing with proper data types."""

import random
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd


class OregonMockDataGenerator:
    """Generate mock Oregon CGT-1 submission data for 2025 template.

    Note: Mock data follows these Oregon-specific requirements:
    - IPA/Contract Name fields are required but values can be blank
    - LOB 7 is only used in TME_ALL sheet
    - LOBs 5 & 6 represent Medicare/Medicaid Dual Eligible (using Paid Amounts)
    - Student health plan members treated as Oregon residents
    - Pharmacy rebates are reported as negative values
    """

    def __init__(self, seed: int = 42, year: int = 2025):
        random.seed(seed)
        self.year = year
        self.providers = self._generate_providers()
        self.line_of_business_codes = {
            1: "Medicare",
            2: "Medicaid",
            3: "Commercial: Full Claims",
            4: "Commercial: Partial Claims",
            5: "Medicare Expenses for Medicare/Medicaid Dual Eligible",
            6: "Medicaid Expenses for Medicare/Medicaid Dual Eligible",
            7: "CCO-F and Medicaid Open Card Carve-Outs",
        }

    def _generate_providers(self, count: int = 10) -> List[Dict]:
        """Generate mock provider data for 2025 template."""
        provider_names = [
            "Oregon Health System",
            "Portland Medical Group",
            "Eugene Healthcare Partners",
            "Salem Community Health",
            "Bend Medical Center",
            "Medford Regional Hospital",
            "Corvallis Health Network",
            "Springfield Medical Associates",
            "Coastal Health Alliance",
            "Central Oregon IPA",
        ]

        providers = []
        for i in range(count):
            # Generate 9-digit TIN with leading zeros - ensure no TIN starts with 0
            # to avoid Excel stripping leading zeros
            tin = f"{random.randint(100000000, 999999999)}"
            provider = {
                "Provider Organization Name": provider_names[i % len(provider_names)]
                + (f" {i//len(provider_names) + 1}" if i >= len(provider_names) else ""),
                "Provider Organization TIN": tin,
                "IPA or Contract Name": "SUB IPA 1" if random.random() > 0.5 else None,
            }
            providers.append(provider)

        return providers

    def generate_cover_page(self, payer_name: str = "Oregon Health Plan") -> pd.DataFrame:
        """Generate Cover Page sheet for 2025 template."""
        # Create a sparse dataframe to match the template structure
        data: list[list[Any]] = [[None] * 4 for _ in range(20)]

        # Add headers and labels
        data[0][0] = "1. Cover Page"
        data[1][0] = "All questions in this tab must be answered."
        data[3][0] = "Contact Information"
        data[4][0] = "Payer Name:"
        data[4][2] = payer_name
        data[5][0] = "Contact Name:"
        data[5][2] = "John Smith"
        data[6][0] = "Contact Email:"
        data[6][2] = "john.smith@oregonhealth.com"

        data[8][0] = "Attestation"
        data[9][0] = (
            "By signing this document, I, as authorized by my organization, hereby declare that the information "
            "in this data submission template is current, complete, correct and true to the best of my knowledge "
            "and belief. I understand that failure to declare that the information is current, complete, correct "
            "and true, will result in an automatic rejection of the data submission. I further acknowledge that "
            "organizations subject to the cost growth target, which exceed the target with statistical confidence "
            "without an acceptable reason, will be subject to accountability mechanisms such as performance "
            "improvement plans (PIP) and financial penalties in accordance with Oregon Revised Statutes, "
            "Chapter 442 — Health Planning, and Oregon Administrative Rules, Chapter 409, Division 65 — "
            "Sustainable Health Care Cost Growth Target Program."
        )
        data[10][0] = "Authorized Signatory Details"
        data[11][0] = "Full Name:"
        data[11][2] = "Jane Doe"
        data[12][0] = "Title/Position:"
        data[12][2] = "Chief Financial Officer"
        data[13][0] = "Email/Contact Information:"
        data[13][2] = "jane.doe@oregonhealth.com"
        data[14][0] = "Signature:"
        data[14][2] = "Jane Doe"
        data[15][0] = "Date:"
        data[15][2] = datetime.now().strftime("%m/%d/%Y")

        return pd.DataFrame(data)

    def generate_tme_all(self) -> pd.DataFrame:
        """Generate TME_ALL sheet for 2025 template."""
        # Headers structure - matching actual template structure
        headers = [
            ["2. Total Medical Expenses: All"] + [None] * 29,  # Row 0: Title
            [None] * 30,  # Row 1: Empty
            ["Black = payer-reported data"] + [None] * 29,  # Row 2
            ["Blue = OHA calculated data"] + [None] * 29,  # Row 3
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,  # Row 4
            [None] * 30,  # Row 5: Empty
            # Row 6: Field codes
            [
                "TMEALL01",
                "TMEALL02",
                "TMEALL03",
                "TMEALL04",
                "TMEALL05",
                "TMEALL06",
                "TMEALL07",
                "TMEALL08",
                "TMEALL09",
                "TMEALL10",
            ]
            + [None] * 20,
            # Row 7: Data types
            [
                "year",
                "code",
                "positive integer",
                "non-negative number",
                "non-negative number",
                "non-negative number",
                "non-negative number",
                "non-negative number",
                "non-negative number",
                "non-negative number",
            ]
            + [None] * 20,
            # Row 8: Column names (this is where validator expects headers)
            [
                "Reporting Year",
                "Line of Business Code",
                "Member Months",
                "Demographic Score",
                "Claims: \nHospital Inpatient",
                "Claims: \nHospital Outpatient",
                "Claims: Professional, Primary Care Providers",
                "Claims: Professional, Specialty Providers",
                "Claims: Professional, Behavior Health Providers",
                "Claims: Professional, Other Providers",
            ]
            + [None] * 20,
        ]

        # Generate data rows
        data_rows = []
        for lob_code in self.line_of_business_codes.keys():
            # Include LOB 7 in TME_ALL as it's allowed only here
            member_months = random.randint(5000, 50000)
            demographic_score = round(random.uniform(0.8, 1.2), 3)

            row = [
                self.year - 1,  # Reporting Year (previous year)
                lob_code,  # Line of Business Code
                member_months,  # Member Months
                demographic_score,  # Demographic Score
                round(member_months * random.uniform(200, 400), 2),  # Hospital Inpatient
                round(member_months * random.uniform(150, 300), 2),  # Hospital Outpatient
                round(member_months * random.uniform(50, 100), 2),  # Primary Care
                round(member_months * random.uniform(100, 200), 2),  # Specialty
                round(member_months * random.uniform(20, 50), 2),  # Behavioral Health
                round(member_months * random.uniform(30, 60), 2),  # Other Providers
            ] + [None] * 20

            data_rows.append(row)

        # Combine headers and data
        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_tme_prov(self) -> pd.DataFrame:
        """Generate TME_PROV sheet for 2025 template."""
        headers = [
            ["3. Total Medical Expenses: Member Months Attributed to Provider Organizations"] + [None] * 29,  # Row 0
            [None] * 30,  # Row 1: Empty
            ["Black = payer-reported data"] + [None] * 29,  # Row 2
            ["Blue = OHA calculated data"] + [None] * 29,  # Row 3
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,  # Row 4
            [None] * 30,  # Row 5: Empty
            [None] * 30,  # Row 6: Empty
            [None] * 30,  # Row 7: Empty
            # Row 8: Field codes
            ["TMEPRV01", "TMEPRV02", "TMEPRV03", "TMEPRV04", "TMEPRV06", "TMEPRV07", "TMEPRV08"] + [None] * 23,
            # Row 9: Data types
            [
                "Year",
                "Code",
                "free text, blank is not allowed",
                "free text, blank allowed",
                "Code",
                "positive integer",
                "non-negative number",
            ]
            + [None] * 23,
            # Row 10: Column names (this is where validator expects headers)
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Attribution Hierarchy Code",
                "Member Months",
                "Demographic Score",
            ]
            + [None] * 23,
        ]

        data_rows = []
        for provider in self.providers:
            for lob_code in [1, 2, 3, 4]:  # Main LOB codes for providers
                # Some providers might not have all LOBs
                if random.random() > 0.2:  # 80% chance of having this LOB
                    member_months = random.randint(15, 5000)  # Some may be below threshold

                    row = [
                        self.year - 1,
                        lob_code,
                        provider["Provider Organization Name"],
                        provider.get("IPA or Contract Name") or "",  # Optional field - use empty if None
                        1,  # Attribution Hierarchy Code (primary)
                        member_months,
                        round(random.uniform(0.8, 1.2), 3),  # Demographic Score
                    ] + [None] * 23

                    data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_tme_unattr(self) -> pd.DataFrame:
        """Generate TME_UNATTR sheet for rolled-up low member months data."""
        headers = [
            ["4. Total Medical Expenses: Member Months Unattributed to Provider Organizations"] + [None] * 29,  # Row 0
            [None] * 30,  # Row 1: Empty
            ["Black = payer-reported data"] + [None] * 29,  # Row 2
            ["Blue = OHA calculated data"] + [None] * 29,  # Row 3
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,  # Row 4
            [None] * 30,  # Row 5: Empty
            # Row 6: Field codes
            ["TMEUNA01", "TMEUNA02", "TMEUNA03", "TMEUNA04"] + [None] * 26,
            # Row 7: Data types
            ["year", "code", "positive integer", "non-negative number"] + [None] * 26,
            # Row 8: Column names (this is where validator expects headers)
            ["Reporting Year", "Line of Business code", "Member Months", "Demographic Score"] + [None] * 26,
        ]

        data_rows = []
        # Add some rolled up data for LOBs that had low member months
        for lob_code in [1, 2, 3, 4]:
            if random.random() > 0.5:  # 50% chance of having unattributed members
                member_months = random.randint(50, 500)  # Rolled up from multiple small providers

                row = [self.year - 1, lob_code, member_months] + [None] * 27
                data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_market_enroll(self) -> pd.DataFrame:
        """Generate MARKET_ENROLL sheet for market-wide enrollment data."""
        headers = [
            ["5. Market Enrollment"] + [None] * 29,  # Row 0
            [None] * 30,  # Row 1: Empty
            ["Black = payer-reported data"] + [None] * 29,  # Row 2
            ["Blue = OHA calculated data"] + [None] * 29,  # Row 3
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,  # Row 4
            [None] * 30,  # Row 5: Empty
            # Row 6: Field codes
            ["MED01", "MED02", "MED03"] + [None] * 27,
            # Row 7: Data types
            [None, "non-negative integer", "non-negative integer"] + [None] * 27,
            # Row 8: Column names (different structure for MARKET_ENROLL)
            ["Market Enrollment Category", f"Year {2023} Member Months", f"Year {2024} Member Months"] + [None] * 27,
        ]

        data_rows = []
        # MARKET_ENROLL has different structure - categories, not LOB codes
        market_categories = [
            "1. Large group (51 + employees), fully insured",
            "2. Small group (2-50 employees), fully insured",
            "3. Individual",
            "4. Medicare",
            "5. Medicaid",
            "6. Self-insured",
            "7. Student health plan",
        ]

        for category in market_categories:
            year_2023_months = random.randint(10000, 100000) if random.random() > 0.3 else None
            year_2024_months = random.randint(10000, 100000) if random.random() > 0.3 else None
            row = [category, year_2023_months, year_2024_months] + [None] * 27
            data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_rx_med_prov(self) -> pd.DataFrame:
        """Generate RX_MED_PROV sheet for prescription and medical data by provider."""
        headers = [
            ["6. Prescription and Medical: Member Months Attributed to Provider Organizations"] + [None] * 29,
            ["Black = payer-reported data"] + [None] * 29,
            ["Blue = OHA calculated data"] + [None] * 29,
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,
            [None] * 30,
            [None] * 30,
            [None] * 30,  # Row 6: Empty
            [None] * 30,  # Row 7: Empty
            # Row 8: Field codes
            ["TMERXPRV01", "TMERXPRV02", "TMERXPRV03", "TMEPRV04", "TMERXPRV05"] + [None] * 25,
            # Row 9: Data types
            [
                "Year",
                "Code",
                "free text, blank is not allowed",
                "free text, blank allowed",
                "Code",
            ]
            + [None] * 25,
            # Row 10: Column names (this is where validator expects headers)
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name",
                "IPA or Contract Name\n(If applicable/available)",
                "Allowed Pharmacy",
                "Net Paid Medical",
            ]
            + [None] * 24,
        ]

        data_rows = []
        for provider in self.providers:
            for lob_code in [1, 2, 3, 4]:
                if random.random() > 0.2:
                    allowed_pharmacy = round(random.uniform(10000, 100000), 2)
                    net_paid_medical = round(random.uniform(50000, 500000), 2)

                    row = [
                        self.year - 1,
                        lob_code,
                        provider["Provider Organization TIN"],
                        provider["Provider Organization Name"],
                        allowed_pharmacy,
                        net_paid_medical,
                    ] + [None] * 24

                    data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_rx_med_unattr(self) -> pd.DataFrame:
        """Generate RX_MED_UNATTR sheet for unattributed prescription and medical data."""
        headers = [
            ["7. Prescription and Medical: Member Months Unattributed to Provider Organizations"] + [None] * 29,
            ["Black = payer-reported data"] + [None] * 29,
            ["Blue = OHA calculated data"] + [None] * 29,
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,
            [None] * 30,
            [None] * 30,
            # Row 6: Field codes (should match actual template)
            ["RXMEDUNA01", "RXMEDUNA02"] + [None] * 28,
            # Row 7: Data types
            ["year", "code"] + [None] * 28,
            # Row 8: Column names (this is where validator expects headers)
            [
                "Reporting Year",
                "Line of Business Code",
            ]
            + [None] * 28,
        ]

        data_rows = []
        for lob_code in [1, 2, 3, 4]:
            if random.random() > 0.5:
                # Simplified data for RX_MED_UNATTR - just year and LOB
                row = [
                    self.year - 1,
                    lob_code,
                ] + [None] * 28

                data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_rx_rebate(self) -> pd.DataFrame:
        """Generate RX_REBATE sheet for prescription rebate data."""
        headers = [
            ["8. Prescription Rebates"] + [None] * 29,  # Row 0
            [None] * 30,  # Row 1: Empty
            ["Black = payer-reported data"] + [None] * 29,  # Row 2
            ["Blue = OHA calculated data"] + [None] * 29,  # Row 3
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 29,  # Row 4
            [None] * 30,  # Row 5: Empty
            # Row 6: Field codes
            ["RXR01", "RXR02", "RX07", "RXR03", "RXR04"] + [None] * 25,
            # Row 7: Data types
            ["year", "Code", "free text, blank is allowed", "non-positive number", "non-positive number"] + [None] * 25,
            # Row 8: Column names (this is where validator expects headers)
            [
                "Reporting Year",
                "Line of Business Code",
                "Provider Organization Name (Optional)",
                "Medical Pharmacy Rebate Amount",
                "Retail Pharmacy Rebate Amount",
            ]
            + [None] * 25,
        ]

        data_rows = []
        # Not all payers have rebates
        if random.random() > 0.3:  # 70% chance of having rebate data
            for lob_code in [1, 2, 3, 4, 5, 6]:  # LOB 7 not allowed in RX_REBATE
                if random.random() > 0.4:  # 60% chance per LOB
                    # Rebates must be negative or zero
                    medical_rebate = -round(random.uniform(1000, 25000), 2)
                    retail_rebate = -round(random.uniform(5000, 50000), 2)
                    total_rebate = medical_rebate + retail_rebate
                    row = [self.year - 1, lob_code, medical_rebate, retail_rebate, total_rebate] + [None] * 25
                    data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_prov_id(self) -> pd.DataFrame:
        """Generate PROV_ID sheet with provider identification data."""
        headers = [
            [None] * 5,
            ["When the payer-entered data does not meet the desired format, the cell will become red."] + [None] * 4,
            [None] * 5,
            [None] * 5,
            [None] * 5,
            [None] * 5,
            # Row 6: Field codes - PRV01, PRV03, PRV02 (note the specific order)
            ["PRV01", "PRV03", "PRV02"] + [None] * 2,
            # Data types
            ["free text", "free text", "text, 9 digits including leading zero"] + [None] * 2,
            # Column names
            [
                "Provider Organization Name",
                "IPA or Contract Name (If applicable/available)",
                "Provider Organization TIN",
            ]
            + [None] * 2,
        ]

        data_rows = []
        for provider in self.providers:
            row = [
                provider["Provider Organization Name"],
                provider.get("IPA or Contract Name") or "",  # Middle column - use empty if None
                provider["Provider Organization TIN"],  # TIN is last
            ] + [None] * 2
            data_rows.append(row)

        all_rows = headers + data_rows
        return pd.DataFrame(all_rows)

    def generate_contents_sheet(self) -> pd.DataFrame:
        """Generate Contents sheet with template information."""
        data = [
            [
                "Oregon's Health Care Cost Growth Target Program - Data Submission Template (CGT-1)",
                "Version 5.0, June 2025",
            ],
            [None, None],
            ["This workbook contains the following tabs:", None],
            [None, None],
            ["Tab", "Description"],
            ["1. Cover Page", "Contact information and attestation"],
            ["2. TME_ALL", "Total Medical Expenses for all members"],
            ["3. TME_PROV", "Total Medical Expenses by provider"],
            ["4. TME_UNATTR", "Total Medical Expenses for unattributed members"],
            ["5. MARKET_ENROLL", "Market-wide enrollment data"],
            ["6. RX_MED_PROV", "Prescription and medical expenses by provider"],
            ["7. RX_MED_UNATTR", "Prescription and medical expenses for unattributed members"],
            ["8. RX_REBATE", "Prescription rebate information"],
            ["9. PROV_ID", "Provider identification information"],
            ["Line of Business Code", "Reference table for LOB codes"],
            ["Attribution Hierarchy Code", "Reference table for attribution codes"],
            ["Demographic Tables", "Reference tables for demographic factors"],
            ["TME Validation", "Automated validation checks"],
            ["RX_MED_PROV Validation", "Automated validation checks"],
            ["RX_MED_UNATTR Validation", "Automated validation checks"],
            ["Provider Check", "Provider data validation"],
        ]

        return pd.DataFrame(data)

    def save_to_excel(self, output_filename: str, include_validations: bool = True) -> Path:
        """Save all sheets to an Excel file matching the 2025 CGT-1 template format."""
        output_path = Path(output_filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Generate all required sheets
        contents_df = self.generate_contents_sheet()
        cover_page_df = self.generate_cover_page()
        tme_all_df = self.generate_tme_all()
        tme_prov_df = self.generate_tme_prov()
        tme_unattr_df = self.generate_tme_unattr()
        market_enroll_df = self.generate_market_enroll()
        rx_med_prov_df = self.generate_rx_med_prov()
        rx_med_unattr_df = self.generate_rx_med_unattr()
        rx_rebate_df = self.generate_rx_rebate()
        prov_id_df = self.generate_prov_id()

        # Write to Excel
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Write sheets in the correct order
            contents_df.to_excel(writer, sheet_name="Contents", index=False, header=False)
            cover_page_df.to_excel(writer, sheet_name="1. Cover Page", index=False, header=False)
            tme_all_df.to_excel(writer, sheet_name="2. TME_ALL", index=False, header=False)
            tme_prov_df.to_excel(writer, sheet_name="3. TME_PROV", index=False, header=False)
            tme_unattr_df.to_excel(writer, sheet_name="4. TME_UNATTR", index=False, header=False)
            market_enroll_df.to_excel(writer, sheet_name="5. MARKET_ENROLL", index=False, header=False)
            rx_med_prov_df.to_excel(writer, sheet_name="6. RX_MED_PROV", index=False, header=False)
            rx_med_unattr_df.to_excel(writer, sheet_name="7. RX_MED_UNATTR", index=False, header=False)
            rx_rebate_df.to_excel(writer, sheet_name="8. RX_REBATE", index=False, header=False)
            prov_id_df.to_excel(writer, sheet_name="9. PROV_ID", index=False, header=False)

            # Add reference sheets
            if include_validations:
                # Line of Business Code reference
                lob_data = [
                    [None, None],
                    ["Line of Business Code", "Description"],
                    [1, "Medicare"],
                    [2, "Medicaid"],
                    [3, "Commercial: Full Claims"],
                    [4, "Commercial: Partial Claims"],
                    [5, "Medicare Expenses for Medicare/Medicaid Dual Eligible"],
                    [6, "Medicaid Expenses for Medicare/Medicaid Dual Eligible"],
                    [
                        7,
                        "CCO-F expenses (for use by CCOs) or Medicaid Carve-Outs (for use by Medicaid FFS) - "
                        "for use in TME_ALL tab only",
                    ],
                ]
                lob_df = pd.DataFrame(lob_data)
                lob_df.to_excel(writer, sheet_name="Line of Business Code", index=False, header=False)

            # Format TIN columns as text in PROV_ID sheet
            workbook = writer.book
            prov_sheet = workbook["9. PROV_ID"]

            # TIN is in column C (3rd column)
            for row in range(10, 10 + len(self.providers)):  # Data starts at row 10
                cell = prov_sheet.cell(row=row, column=3)
                cell.number_format = "@"  # Text format

        print(f"Mock Oregon CGT-1 2025 template data saved to: {output_path}")
        return output_path


def generate_mock_submission(
    output_path: str = "./mock_data/oregon/oregon_2025_submission.xlsx",
    include_validations: bool = True,
    year: int = 2025,
) -> Path:
    """Generate a mock Oregon CGT-1 2025 submission file with valid data."""
    generator = OregonMockDataGenerator(year=year)
    return generator.save_to_excel(output_path, include_validations)


def generate_fail_submission(
    output_path: str = "./mock_data/oregon/oregon_2025_submission_fail.xlsx", year: int = 2025
) -> Path:
    """Generate a mock Oregon CGT-1 2025 submission file with validation errors."""
    generator = OregonMockDataGenerator(year=year)

    # Modify some data to create validation errors
    # Remove some required providers
    generator.providers = generator.providers[:3]

    # Generate the file
    path = generator.save_to_excel(output_path)

    # Post-process to introduce errors
    import openpyxl

    wb = openpyxl.load_workbook(path)

    # Clear some required fields in Cover Page
    cover_sheet = wb["1. Cover Page"]
    cover_sheet.cell(row=5, column=3).value = "[Input Required]"  # Clear Payer Name
    cover_sheet.cell(row=12, column=3).value = ""  # Clear Full Name
    cover_sheet.cell(row=15, column=3).value = "[Input Required]"  # Clear Signature

    # Add invalid Line of Business codes
    tme_all_sheet = wb["2. TME_ALL"]
    if tme_all_sheet.cell(row=11, column=2).value:
        tme_all_sheet.cell(row=11, column=2).value = 99  # Invalid LOB code

    # Add LOB 7 in non-TME_ALL sheet (invalid)
    tme_prov_sheet = wb["3. TME_PROV"]
    if tme_prov_sheet.cell(row=13, column=2).value:
        tme_prov_sheet.cell(row=13, column=2).value = 7  # LOB 7 not allowed here

    # Add invalid TIN format
    prov_id_sheet = wb["9. PROV_ID"]
    if prov_id_sheet.cell(row=9, column=3).value:
        prov_id_sheet.cell(row=9, column=3).value = "12345"  # Too short TIN

    # Add positive pharmacy rebate (invalid)
    rx_rebate_sheet = wb["8. RX_REBATE"]
    if rx_rebate_sheet.cell(row=11, column=3).value:
        rx_rebate_sheet.cell(row=11, column=3).value = 10000  # Positive rebate is invalid

    wb.save(path)
    return path


if __name__ == "__main__":
    # Generate valid mock data
    valid_file = generate_mock_submission()
    print(f"Successfully generated valid mock Oregon CGT-1 2025 submission: {valid_file}")

    # Generate fail mock data
    fail_file = generate_fail_submission()
    print(f"Successfully generated fail mock Oregon CGT-1 2025 submission: {fail_file}")
